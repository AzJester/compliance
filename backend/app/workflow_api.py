from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from collections.abc import Iterable
from dataclasses import dataclass
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from fastapi.encoders import jsonable_encoder
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_, select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from .database import get_session
from .freshness import (
    invalidate_requirement_extraction,
    mark_crosswalk_stale,
    reset_package_verification,
)
from .models import (
    CDRL,
    SOLICITATION_DOCUMENT_CLASSIFICATIONS,
    ActionStatus,
    CDRLAdjudication,
    CDRLAdjudicationStatus,
    CrosswalkFinding,
    CrosswalkRunState,
    CrosswalkStatus,
    Document,
    DocumentClassification,
    DocumentProfile,
    DocumentStatus,
    IntakeVerification,
    IntakeVerificationStatus,
    Project,
    ProjectAction,
    ProjectWorkflow,
    ProposalEvidence,
    Requirement,
    RequirementExtractionState,
    RequirementSection,
    ValidationStatus,
    WorkflowStage,
    WorkflowStatus,
    utc_now,
)
from .requirements_rules import RULE_VERSION
from .requirements_service import cdrl_response
from .schemas import (
    CDRLAdjudicationPatch,
    CDRLAdjudicationResponse,
    CrosswalkFindingPatch,
    CrosswalkFindingResponse,
    CrosswalkGenerateSummary,
    DocumentProfilePatch,
    DocumentResponse,
    DocumentTextResponse,
    IntakeVerificationCreate,
    IntakeVerificationPatch,
    IntakeVerificationResponse,
    ProjectActionCreate,
    ProjectActionPatch,
    ProjectActionResponse,
    ProjectPatch,
    ProjectResponse,
    ProposalEvidenceCreate,
    ProposalEvidenceResponse,
    ReadinessResponse,
    StageProgressResponse,
    WorkflowPatch,
    WorkflowResponse,
)

router = APIRouter(prefix="/api/projects/{project_id}", tags=["workflow"])

_CDRL_SNAPSHOT_FIELDS = (
    "fingerprint",
    "source_text",
    "source_start",
    "source_end",
    "source_locator",
    "rule_version",
    "incomplete",
    "missing_fields",
    "source_truncated",
    "block_a_contract_line_item_number",
    "block_b_exhibit",
    "block_c_category",
    "block_d_system_item",
    "block_e_contract_pr_number",
    "block_f_contractor",
    "block_1_data_item_number",
    "block_2_title",
    "block_3_subtitle",
    "block_4_authority",
    "block_5_contract_reference",
    "block_6_requiring_office",
    "block_7_dd250_requirement",
    "block_8_approval_code",
    "block_9_distribution_statement",
    "block_10_frequency",
    "block_11_as_of_date",
    "block_12_first_submission",
    "block_13_subsequent_submission",
    "block_14_distribution",
    "block_15_total",
    "block_16_remarks",
    "block_17_price_group",
    "block_18_estimated_total_price",
)

_STOP_WORDS = {
    "a",
    "an",
    "and",
    "are",
    "as",
    "at",
    "be",
    "by",
    "for",
    "from",
    "in",
    "is",
    "it",
    "of",
    "on",
    "or",
    "shall",
    "should",
    "that",
    "the",
    "this",
    "to",
    "will",
    "with",
}
_TOKEN_PATTERN = re.compile(r"[a-z0-9]+(?:[./%-][a-z0-9]+)*", re.IGNORECASE)
_NUMBER_PATTERN = re.compile(r"\b\d+(?:\.\d+)?(?:%|[/-]\d+)*\b")
_DEFAULT_INTAKE_CHECKS = (
    ("solicitation_identified", "Base solicitation identified"),
    ("amendments_accounted", "All amendments accounted for"),
    ("attachments_inventoried", "Attachments inventoried"),
    ("q_and_a_incorporated", "Questions and answers incorporated"),
    ("file_errors_resolved", "Unreadable, duplicate, and OCR issues resolved"),
    ("documents_classified", "All documents classified"),
)


@dataclass(frozen=True, slots=True)
class _ProposalChunk:
    document: Document
    start: int
    end: int
    excerpt: str
    tokens: frozenset[str]


def _project(session: Session, project_id: str) -> Project:
    project = session.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found.")
    return project


def _document(session: Session, project_id: str, document_id: str) -> Document:
    document = session.scalar(
        select(Document)
        .options(selectinload(Document.blob), selectinload(Document.workflow_profile))
        .where(Document.id == document_id, Document.project_id == project_id)
    )
    if document is None:
        raise HTTPException(status_code=404, detail="Document not found.")
    return document


def _cdrl(session: Session, project_id: str, cdrl_id: str) -> CDRL:
    cdrl = session.scalar(
        select(CDRL)
        .options(
            selectinload(CDRL.document).selectinload(Document.workflow_profile),
            selectinload(CDRL.adjudication),
        )
        .where(CDRL.id == cdrl_id, CDRL.project_id == project_id)
    )
    if cdrl is None:
        raise HTTPException(status_code=404, detail="CDRL not found.")
    return cdrl


def _cdrl_source_fingerprint(cdrl: CDRL) -> str:
    snapshot = {
        field: sorted(value) if field == "missing_fields" and value else value
        for field in _CDRL_SNAPSHOT_FIELDS
        if (value := getattr(cdrl, field)) is not None
    }
    snapshot.update(
        {
            "document_blob_sha256": cdrl.document.blob_sha256,
            "document_classification": cdrl.document.classification.value,
            "document_text_sha256": hashlib.sha256(
                cdrl.document.extracted_text.encode("utf-8")
            ).hexdigest(),
        }
    )
    encoded = json.dumps(snapshot, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _cdrl_is_incomplete(cdrl: CDRL) -> bool:
    return bool(cdrl.incomplete or cdrl.missing_fields or cdrl.source_truncated)


def _cdrl_adjudication_response(cdrl: CDRL) -> CDRLAdjudicationResponse:
    adjudication = cdrl.adjudication
    current_fingerprint = _cdrl_source_fingerprint(cdrl)
    context_only = cdrl.document.classification == DocumentClassification.REFERENCE
    fresh = bool(
        adjudication is not None
        and adjudication.source_fingerprint is not None
        and adjudication.source_fingerprint == current_fingerprint
    )
    incomplete = _cdrl_is_incomplete(cdrl)
    status_value = (
        adjudication.status if adjudication is not None else CDRLAdjudicationStatus.PENDING
    )
    valid_waiver = bool(
        fresh
        and status_value == CDRLAdjudicationStatus.WAIVED
        and adjudication is not None
        and adjudication.reviewer
        and adjudication.waiver_reason
    )
    valid_review = bool(
        fresh
        and not incomplete
        and status_value == CDRLAdjudicationStatus.REVIEWED
        and adjudication is not None
        and adjudication.reviewer
    )
    return CDRLAdjudicationResponse(
        cdrl_id=cdrl.id,
        project_id=cdrl.project_id,
        status=status_value,
        reviewer=adjudication.reviewer if adjudication is not None else None,
        waiver_reason=adjudication.waiver_reason if adjudication is not None else None,
        reviewed_at=adjudication.reviewed_at if adjudication is not None else None,
        updated_at=adjudication.updated_at if adjudication is not None else None,
        source_fingerprint=(adjudication.source_fingerprint if adjudication is not None else None),
        fresh=fresh,
        context_only=context_only,
        incomplete=incomplete,
        missing_fields=list(cdrl.missing_fields or []),
        effective_ready=context_only or valid_waiver or valid_review,
    )


def _workflow(session: Session, project_id: str) -> ProjectWorkflow:
    workflow = session.get(ProjectWorkflow, project_id)
    if workflow is None:
        workflow = ProjectWorkflow(project_id=project_id)
        session.add(workflow)
        session.flush()
    return workflow


def _finding(session: Session, project_id: str, finding_id: str) -> CrosswalkFinding:
    finding = session.scalar(
        select(CrosswalkFinding)
        .join(Requirement, Requirement.id == CrosswalkFinding.requirement_id)
        .join(Document, Document.id == Requirement.document_id)
        .outerjoin(DocumentProfile, DocumentProfile.document_id == Document.id)
        .options(
            selectinload(CrosswalkFinding.requirement),
            selectinload(CrosswalkFinding.evidence)
            .selectinload(ProposalEvidence.document)
            .selectinload(Document.workflow_profile),
        )
        .where(
            CrosswalkFinding.id == finding_id,
            CrosswalkFinding.project_id == project_id,
            or_(
                DocumentProfile.document_id.is_(None),
                DocumentProfile.classification.in_(SOLICITATION_DOCUMENT_CLASSIFICATIONS),
            ),
        )
    )
    if finding is None:
        raise HTTPException(status_code=404, detail="Crosswalk finding not found.")
    return finding


def _finding_attention_reasons(finding: CrosswalkFinding) -> list[str]:
    reasons: list[str] = []
    if finding.stale:
        reasons.append("Reanalyze this finding after source changes.")
    if any(
        evidence.document.classification != DocumentClassification.PROPOSAL_VOLUME
        for evidence in finding.evidence
    ):
        reasons.append("Replace evidence that does not come from a proposal volume.")
    if (
        finding.status
        in {CrosswalkStatus.COVERED, CrosswalkStatus.PARTIAL, CrosswalkStatus.CONFLICT}
        and not finding.evidence
    ):
        reasons.append("Add current proposal evidence for this finding.")
    if finding.status != finding.candidate_status and not finding.human_verified:
        reasons.append("Confirm or remove the manual status override.")
    if finding.status in {
        CrosswalkStatus.PARTIAL,
        CrosswalkStatus.MISSING,
        CrosswalkStatus.CONFLICT,
    }:
        reasons.append(f"Resolve the {finding.status.value.lower()} proposal coverage result.")
    return reasons


def _finding_response(finding: CrosswalkFinding) -> CrosswalkFindingResponse:
    reasons = _finding_attention_reasons(finding)
    return CrosswalkFindingResponse.model_validate(finding).model_copy(
        update={"needs_attention": bool(reasons), "attention_reasons": reasons}
    )


@router.patch("", response_model=ProjectResponse)
def patch_project(
    project_id: str,
    patch: ProjectPatch,
    session: Session = Depends(get_session),
) -> Project:
    project = _project(session, project_id)
    updates = patch.model_dump(exclude_unset=True)
    effective_due = updates.get("due_at", project.due_at)
    effective_timezone = updates.get("due_timezone", project.due_timezone)
    if effective_due is not None and effective_timezone is None:
        raise HTTPException(
            status_code=422,
            detail="due_timezone is required when due_at is provided",
        )
    if updates.get("due_at", ...) is None and "due_timezone" not in updates:
        updates["due_timezone"] = None
    for field, value in updates.items():
        setattr(project, field, value)
    session.commit()
    session.refresh(project)
    return project


@router.get("/workflow", response_model=WorkflowResponse)
def get_workflow(project_id: str, session: Session = Depends(get_session)) -> ProjectWorkflow:
    _project(session, project_id)
    workflow = _workflow(session, project_id)
    session.commit()
    session.refresh(workflow)
    return workflow


@router.patch("/workflow", response_model=WorkflowResponse)
def patch_workflow(
    project_id: str,
    patch: WorkflowPatch,
    session: Session = Depends(get_session),
) -> ProjectWorkflow:
    _project(session, project_id)
    workflow = _workflow(session, project_id)
    for field, value in patch.model_dump(exclude_unset=True).items():
        setattr(workflow, field, value)
    if workflow.status != WorkflowStatus.BLOCKED:
        workflow.blocker_summary = (
            patch.blocker_summary if "blocker_summary" in patch.model_fields_set else None
        )
    session.commit()
    session.refresh(workflow)
    return workflow


@router.patch("/documents/{document_id}/profile", response_model=DocumentResponse)
def patch_document_profile(
    project_id: str,
    document_id: str,
    patch: DocumentProfilePatch,
    session: Session = Depends(get_session),
) -> Document:
    _project(session, project_id)
    document = _document(session, project_id, document_id)
    previous_classification = document.classification
    if patch.classification == DocumentClassification.PROPOSAL_VOLUME:
        requirement_count = session.scalar(
            select(func.count(Requirement.id)).where(Requirement.document_id == document_id)
        )
        if requirement_count:
            raise HTTPException(
                status_code=409,
                detail=(
                    "This document already produced solicitation requirements. Classify proposal "
                    "volumes before running requirement extraction."
                ),
            )

    profile = document.workflow_profile
    if profile is None:
        profile = DocumentProfile(document_id=document.id, project_id=project_id)
        session.add(profile)
        document.workflow_profile = profile
    profile.classification = patch.classification
    profile.volume_name = patch.volume_name
    profile.notes = patch.classification_notes
    if previous_classification != patch.classification:
        invalidate_requirement_extraction(session, document.id)
        session.execute(
            update(CDRLAdjudication)
            .where(
                CDRLAdjudication.cdrl_id.in_(select(CDRL.id).where(CDRL.document_id == document.id))
            )
            .values(source_fingerprint=None, updated_at=utc_now())
        )
        changed_classifications = {
            previous_classification,
            patch.classification,
        }
        if changed_classifications & SOLICITATION_DOCUMENT_CLASSIFICATIONS:
            reset_package_verification(session, project_id)
        if changed_classifications & (
            SOLICITATION_DOCUMENT_CLASSIFICATIONS | {DocumentClassification.PROPOSAL_VOLUME}
        ):
            mark_crosswalk_stale(session, project_id)
    session.commit()
    session.refresh(profile)
    return document


@router.get("/documents/{document_id}/text", response_model=DocumentTextResponse)
def get_document_text(
    project_id: str,
    document_id: str,
    start: int = Query(default=0, ge=0),
    limit: int = Query(default=12_000, ge=1, le=20_000),
    session: Session = Depends(get_session),
) -> DocumentTextResponse:
    _project(session, project_id)
    document = _document(session, project_id, document_id)
    total = len(document.extracted_text)
    bounded_start = min(start, total)
    end = min(bounded_start + limit, total)
    return DocumentTextResponse(
        document_id=document.id,
        name=document.name,
        total_characters=total,
        start=bounded_start,
        end=end,
        text=document.extracted_text[bounded_start:end],
        truncated=end < total,
    )


@router.get("/cdrl-adjudications", response_model=list[CDRLAdjudicationResponse])
def list_cdrl_adjudications(
    project_id: str, session: Session = Depends(get_session)
) -> list[CDRLAdjudicationResponse]:
    _project(session, project_id)
    cdrls = session.scalars(
        select(CDRL)
        .options(
            selectinload(CDRL.document).selectinload(Document.workflow_profile),
            selectinload(CDRL.adjudication),
        )
        .where(CDRL.project_id == project_id)
        .order_by(CDRL.document_id, CDRL.source_start, CDRL.id)
    )
    return [_cdrl_adjudication_response(cdrl) for cdrl in cdrls]


@router.put(
    "/cdrls/{cdrl_id}/adjudication",
    response_model=CDRLAdjudicationResponse,
)
def put_cdrl_adjudication(
    project_id: str,
    cdrl_id: str,
    patch: CDRLAdjudicationPatch,
    session: Session = Depends(get_session),
) -> CDRLAdjudicationResponse:
    _project(session, project_id)
    cdrl = _cdrl(session, project_id, cdrl_id)
    adjudication = cdrl.adjudication
    if patch.expected_updated_at is not None and (
        adjudication is None or adjudication.updated_at != patch.expected_updated_at
    ):
        raise HTTPException(
            status_code=409,
            detail="The CDRL adjudication changed after it was loaded. Refresh and try again.",
        )
    if adjudication is None:
        adjudication = CDRLAdjudication(cdrl_id=cdrl.id, project_id=project_id)
        cdrl.adjudication = adjudication
        session.add(adjudication)

    adjudication.status = patch.status
    if patch.status == CDRLAdjudicationStatus.PENDING:
        adjudication.reviewer = None
        adjudication.waiver_reason = None
        adjudication.source_fingerprint = None
        adjudication.reviewed_at = None
    else:
        adjudication.reviewer = patch.reviewer
        adjudication.waiver_reason = (
            patch.waiver_reason if patch.status == CDRLAdjudicationStatus.WAIVED else None
        )
        adjudication.source_fingerprint = _cdrl_source_fingerprint(cdrl)
        adjudication.reviewed_at = utc_now()
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        raise HTTPException(
            status_code=409,
            detail="The CDRL adjudication changed concurrently. Refresh and try again.",
        ) from None
    return _cdrl_adjudication_response(_cdrl(session, project_id, cdrl_id))


@router.post(
    "/intake-verifications/initialize",
    response_model=list[IntakeVerificationResponse],
)
def initialize_intake_verifications(
    project_id: str, session: Session = Depends(get_session)
) -> list[IntakeVerification]:
    _project(session, project_id)
    existing_keys = set(
        session.scalars(
            select(IntakeVerification.check_key).where(
                IntakeVerification.project_id == project_id,
                IntakeVerification.scope == "PROJECT",
            )
        )
    )
    for check_key, label in _DEFAULT_INTAKE_CHECKS:
        if check_key not in existing_keys:
            session.add(
                IntakeVerification(
                    project_id=project_id,
                    scope="PROJECT",
                    check_key=check_key,
                    label=label,
                )
            )
    try:
        session.commit()
    except IntegrityError:
        # Idempotent initialization may race with another request.
        session.rollback()
    return list(
        session.scalars(
            select(IntakeVerification)
            .where(IntakeVerification.project_id == project_id)
            .order_by(IntakeVerification.created_at, IntakeVerification.id)
        )
    )


@router.get("/intake-verifications", response_model=list[IntakeVerificationResponse])
def list_intake_verifications(
    project_id: str, session: Session = Depends(get_session)
) -> list[IntakeVerification]:
    _project(session, project_id)
    return list(
        session.scalars(
            select(IntakeVerification)
            .where(IntakeVerification.project_id == project_id)
            .order_by(IntakeVerification.created_at, IntakeVerification.id)
        )
    )


@router.post(
    "/intake-verifications",
    response_model=IntakeVerificationResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_intake_verification(
    project_id: str,
    payload: IntakeVerificationCreate,
    session: Session = Depends(get_session),
) -> IntakeVerification:
    _project(session, project_id)
    if payload.document_id is not None:
        _document(session, project_id, payload.document_id)
    item = IntakeVerification(
        project_id=project_id,
        scope=payload.document_id or "PROJECT",
        **payload.model_dump(),
    )
    session.add(item)
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        raise HTTPException(
            status_code=409,
            detail="That verification check already exists for this scope.",
        ) from None
    session.refresh(item)
    return item


@router.patch("/intake-verifications/{verification_id}", response_model=IntakeVerificationResponse)
def patch_intake_verification(
    project_id: str,
    verification_id: str,
    patch: IntakeVerificationPatch,
    session: Session = Depends(get_session),
) -> IntakeVerification:
    _project(session, project_id)
    item = session.scalar(
        select(IntakeVerification).where(
            IntakeVerification.id == verification_id,
            IntakeVerification.project_id == project_id,
        )
    )
    if item is None:
        raise HTTPException(status_code=404, detail="Intake verification not found.")
    for field, value in patch.model_dump(exclude_unset=True).items():
        setattr(item, field, value)
    session.commit()
    session.refresh(item)
    return item


def _validate_action_links(
    session: Session,
    project_id: str,
    requirement_id: str | None,
    finding_id: str | None,
) -> None:
    if requirement_id is not None:
        requirement = session.scalar(
            select(Requirement.id).where(
                Requirement.id == requirement_id,
                Requirement.project_id == project_id,
            )
        )
        if requirement is None:
            raise HTTPException(status_code=404, detail="Requirement not found.")
    if finding_id is not None:
        finding = session.scalar(
            select(CrosswalkFinding.id).where(
                CrosswalkFinding.id == finding_id,
                CrosswalkFinding.project_id == project_id,
            )
        )
        if finding is None:
            raise HTTPException(status_code=404, detail="Crosswalk finding not found.")


@router.get("/actions", response_model=list[ProjectActionResponse])
def list_actions(project_id: str, session: Session = Depends(get_session)) -> list[ProjectAction]:
    _project(session, project_id)
    return list(
        session.scalars(
            select(ProjectAction)
            .where(ProjectAction.project_id == project_id)
            .order_by(ProjectAction.status, ProjectAction.due_at, ProjectAction.created_at)
        )
    )


@router.post("/actions", response_model=ProjectActionResponse, status_code=status.HTTP_201_CREATED)
def create_action(
    project_id: str,
    payload: ProjectActionCreate,
    session: Session = Depends(get_session),
) -> ProjectAction:
    _project(session, project_id)
    _validate_action_links(session, project_id, payload.requirement_id, payload.finding_id)
    action = ProjectAction(project_id=project_id, **payload.model_dump())
    session.add(action)
    session.commit()
    session.refresh(action)
    return action


@router.patch("/actions/{action_id}", response_model=ProjectActionResponse)
def patch_action(
    project_id: str,
    action_id: str,
    patch: ProjectActionPatch,
    session: Session = Depends(get_session),
) -> ProjectAction:
    _project(session, project_id)
    action = session.scalar(
        select(ProjectAction).where(
            ProjectAction.id == action_id,
            ProjectAction.project_id == project_id,
        )
    )
    if action is None:
        raise HTTPException(status_code=404, detail="Project action not found.")
    for field, value in patch.model_dump(exclude_unset=True).items():
        setattr(action, field, value)
    session.commit()
    session.refresh(action)
    return action


def _tokens(text: str) -> set[str]:
    return {
        token.lower()
        for token in _TOKEN_PATTERN.findall(text)
        if len(token) > 2 and token.lower() not in _STOP_WORDS
    }


def _chunks(document: Document) -> Iterable[tuple[int, int, str]]:
    spans: list[tuple[int, int, str]] = []
    for match in re.finditer(r"\S[^\n]*", document.extracted_text):
        raw = match.group(0).strip()
        if not raw:
            continue
        if len(raw) <= 4_000:
            leading = len(match.group(0)) - len(match.group(0).lstrip())
            start = match.start() + leading
            spans.append((start, start + len(raw), raw))
            continue
        for offset in range(0, len(raw), 3_200):
            excerpt = raw[offset : offset + 4_000]
            spans.append((match.start() + offset, match.start() + offset + len(excerpt), excerpt))
    for index, span in enumerate(spans):
        yield span
        start, end, _ = span
        for following in spans[index + 1 : index + 4]:
            gap = document.extracted_text[end : following[0]]
            if gap.count("\n") > 2:
                break
            end = following[1]
            if end - start > 4_000:
                break
            yield start, end, document.extracted_text[start:end]


def _proposal_chunks(documents: list[Document]) -> list[_ProposalChunk]:
    return [
        _ProposalChunk(
            document=document,
            start=start,
            end=end,
            excerpt=excerpt,
            tokens=frozenset(_tokens(excerpt)),
        )
        for document in documents
        for start, end, excerpt in _chunks(document)
    ]


def _match_requirement(
    requirement: Requirement, proposal_chunks: list[_ProposalChunk]
) -> tuple[CrosswalkStatus, float, Document | None, int, int, str]:
    requirement_tokens = _tokens(requirement.requirement_text)
    if not requirement_tokens:
        return CrosswalkStatus.MISSING, 0.0, None, 0, 0, ""
    best: tuple[float, Document | None, int, int, str] = (0.0, None, 0, 0, "")
    for chunk in proposal_chunks:
        if not chunk.tokens:
            continue
        overlap = requirement_tokens.intersection(chunk.tokens)
        recall = len(overlap) / len(requirement_tokens)
        union = requirement_tokens.union(chunk.tokens)
        jaccard = len(overlap) / len(union)
        score = round((0.75 * recall) + (0.25 * jaccard), 4)
        if score > best[0]:
            best = (score, chunk.document, chunk.start, chunk.end, chunk.excerpt)

    score, document, start, end, excerpt = best
    if document is None or score < 0.18:
        return CrosswalkStatus.MISSING, score, None, 0, 0, ""

    required_numbers = set(_NUMBER_PATTERN.findall(requirement.requirement_text.lower()))
    evidence_numbers = set(_NUMBER_PATTERN.findall(excerpt.lower()))
    if (
        score >= 0.25
        and required_numbers
        and evidence_numbers
        and not required_numbers.issubset(evidence_numbers)
    ):
        candidate_status = CrosswalkStatus.CONFLICT
    elif score >= 0.62:
        candidate_status = CrosswalkStatus.COVERED
    elif score >= 0.25:
        candidate_status = CrosswalkStatus.PARTIAL
    else:
        candidate_status = CrosswalkStatus.MISSING
        document = None
        start = 0
        end = 0
        excerpt = ""
    return candidate_status, score, document, start, end, excerpt


def _candidate_signature(
    requirement: Requirement,
    candidate_status: CrosswalkStatus,
    score: float,
    document: Document | None,
    start: int,
    end: int,
    excerpt: str,
) -> str:
    source = (
        f"{requirement.requirement_text}|{candidate_status.value}|{score}|"
        f"{document.id if document else ''}|{start}|{end}|{excerpt}"
    )
    return hashlib.sha256(source.encode("utf-8")).hexdigest()


def _requirement_input_signature(requirements: Iterable[Requirement]) -> str:
    records = sorted(
        (
            requirement.id,
            requirement.requirement_text,
            requirement.section.value,
            requirement.category.value,
            requirement.obligation_owner.value,
            requirement.applicability.value,
        )
        for requirement in requirements
    )
    payload = json.dumps(records, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _proposal_input_signature(documents: Iterable[Document]) -> str:
    records = sorted(
        (
            document.id,
            document.blob_sha256,
            document.status.value,
            hashlib.sha256(document.extracted_text.encode("utf-8")).hexdigest(),
        )
        for document in documents
    )
    payload = json.dumps(records, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


@router.post("/crosswalk/generate", response_model=CrosswalkGenerateSummary)
def generate_crosswalk(
    project_id: str, session: Session = Depends(get_session)
) -> CrosswalkGenerateSummary:
    _project(session, project_id)
    requirements = list(
        session.scalars(
            select(Requirement)
            .join(Document, Document.id == Requirement.document_id)
            .outerjoin(DocumentProfile, DocumentProfile.document_id == Document.id)
            .where(
                Requirement.project_id == project_id,
                Requirement.validation_status != ValidationStatus.DISMISSED,
                or_(
                    DocumentProfile.document_id.is_(None),
                    DocumentProfile.classification.in_(SOLICITATION_DOCUMENT_CLASSIFICATIONS),
                ),
            )
            .order_by(Requirement.document_id, Requirement.source_start, Requirement.id)
        )
    )
    all_proposal_documents = list(
        session.scalars(
            select(Document)
            .join(DocumentProfile, DocumentProfile.document_id == Document.id)
            .where(
                Document.project_id == project_id,
                DocumentProfile.classification == DocumentClassification.PROPOSAL_VOLUME,
            )
            .order_by(Document.created_at, Document.id)
        )
    )
    proposal_documents = [
        document
        for document in all_proposal_documents
        if document.status == DocumentStatus.EXTRACTED
    ]
    deduplicated_documents: list[Document] = []
    seen_content: set[tuple[str, str]] = set()
    for document in proposal_documents:
        content_key = (document.blob_sha256, document.content_type)
        if content_key not in seen_content:
            deduplicated_documents.append(document)
            seen_content.add(content_key)
    proposal_chunks = _proposal_chunks(deduplicated_documents)

    existing = {
        finding.requirement_id: finding
        for finding in session.scalars(
            select(CrosswalkFinding)
            .options(selectinload(CrosswalkFinding.evidence))
            .where(CrosswalkFinding.project_id == project_id)
        )
    }
    created = 0
    updated = 0
    marked_stale = 0
    generated_at = utc_now()
    for requirement in requirements:
        candidate_status, score, document, start, end, excerpt = _match_requirement(
            requirement, proposal_chunks
        )
        signature = _candidate_signature(
            requirement,
            candidate_status,
            score,
            document,
            start,
            end,
            excerpt,
        )
        finding = existing.get(requirement.id)
        if finding is None:
            finding = CrosswalkFinding(
                project_id=project_id,
                requirement_id=requirement.id,
                candidate_status=candidate_status,
                status=candidate_status,
                score=score,
                candidate_signature=signature,
                generated_at=generated_at,
            )
            session.add(finding)
            session.flush()
            created += 1
        else:
            changed = finding.candidate_signature != signature
            was_stale = finding.stale
            active_manual_override = (
                finding.human_verified and finding.status != finding.candidate_status
            )
            if active_manual_override:
                if changed:
                    finding.stale = True
                    marked_stale += 1
            else:
                finding.status = candidate_status
                finding.stale = False
                if finding.human_verified and (changed or was_stale):
                    finding.human_verified = False
                    finding.reviewer = None
                    finding.reviewed_at = None
            finding.candidate_status = candidate_status
            finding.score = score
            finding.candidate_signature = signature
            finding.generated_at = generated_at
            for evidence in list(finding.evidence):
                if not evidence.is_manual:
                    session.delete(evidence)
            updated += 1

        if document is not None and excerpt:
            session.add(
                ProposalEvidence(
                    project_id=project_id,
                    finding_id=finding.id,
                    document_id=document.id,
                    source_start=start,
                    source_end=end,
                    source_locator=f"characters {start}-{end}",
                    excerpt=excerpt,
                    score=score,
                    is_manual=False,
                )
            )
    run_state = session.get(CrosswalkRunState, project_id)
    if run_state is None:
        run_state = CrosswalkRunState(
            project_id=project_id,
            requirement_signature="",
            proposal_signature="",
            generated_at=generated_at,
        )
        session.add(run_state)
    run_state.requirement_signature = _requirement_input_signature(requirements)
    run_state.proposal_signature = _proposal_input_signature(all_proposal_documents)
    run_state.generated_at = generated_at
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        # A simultaneous generation request may have completed the same unique
        # requirement findings first. Treat that completed register as success.
        completed = session.scalar(
            select(func.count(CrosswalkFinding.id)).where(
                CrosswalkFinding.project_id == project_id,
                CrosswalkFinding.requirement_id.in_([item.id for item in requirements])
                if requirements
                else CrosswalkFinding.id == "",
            )
        )
        if completed != len(requirements):
            raise
        return CrosswalkGenerateSummary(
            requirements_analyzed=len(requirements),
            proposal_documents_analyzed=len(deduplicated_documents),
            findings_created=0,
            findings_updated=len(requirements),
            verified_findings_marked_stale=0,
        )
    return CrosswalkGenerateSummary(
        requirements_analyzed=len(requirements),
        proposal_documents_analyzed=len(deduplicated_documents),
        findings_created=created,
        findings_updated=updated,
        verified_findings_marked_stale=marked_stale,
    )


@router.get("/crosswalk", response_model=list[CrosswalkFindingResponse])
def list_crosswalk(
    project_id: str,
    finding_status: CrosswalkStatus | None = Query(default=None, alias="status"),
    human_verified: bool | None = Query(default=None),
    session: Session = Depends(get_session),
) -> list[CrosswalkFindingResponse]:
    _project(session, project_id)
    statement = (
        select(CrosswalkFinding)
        .join(Requirement, Requirement.id == CrosswalkFinding.requirement_id)
        .join(Document, Document.id == Requirement.document_id)
        .outerjoin(DocumentProfile, DocumentProfile.document_id == Document.id)
        .options(
            selectinload(CrosswalkFinding.requirement),
            selectinload(CrosswalkFinding.evidence)
            .selectinload(ProposalEvidence.document)
            .selectinload(Document.workflow_profile),
        )
        .where(CrosswalkFinding.project_id == project_id)
        .where(Requirement.validation_status != ValidationStatus.DISMISSED)
        .where(
            or_(
                DocumentProfile.document_id.is_(None),
                DocumentProfile.classification.in_(SOLICITATION_DOCUMENT_CLASSIFICATIONS),
            )
        )
    )
    if finding_status is not None:
        statement = statement.where(CrosswalkFinding.status == finding_status)
    if human_verified is not None:
        statement = statement.where(CrosswalkFinding.human_verified == human_verified)
    findings = list(
        session.scalars(
            statement.order_by(
                Requirement.section,
                Requirement.document_id,
                Requirement.source_start,
                CrosswalkFinding.id,
            )
        )
    )
    return [_finding_response(finding) for finding in findings]


@router.patch("/crosswalk/{finding_id}", response_model=CrosswalkFindingResponse)
def patch_crosswalk_finding(
    project_id: str,
    finding_id: str,
    patch: CrosswalkFindingPatch,
    session: Session = Depends(get_session),
) -> CrosswalkFindingResponse:
    _project(session, project_id)
    finding = _finding(session, project_id, finding_id)
    if patch.expected_updated_at is not None and finding.updated_at != patch.expected_updated_at:
        raise HTTPException(
            status_code=409,
            detail="The crosswalk finding changed after it was loaded. Refresh and try again.",
        )
    if finding.human_verified and "status" in patch.model_fields_set and not patch.reviewer:
        raise HTTPException(
            status_code=422,
            detail="reviewer is required to change a verified finding",
        )
    target_status = patch.status if "status" in patch.model_fields_set else finding.status
    target_human_verified = (
        patch.human_verified
        if "human_verified" in patch.model_fields_set
        else finding.human_verified
    )
    target_reviewer = patch.reviewer if "reviewer" in patch.model_fields_set else finding.reviewer
    if target_status != finding.candidate_status and (
        not target_human_verified or not target_reviewer
    ):
        raise HTTPException(
            status_code=422,
            detail="Manual status overrides require human verification and a reviewer.",
        )
    if (
        patch.human_verified is True
        and target_status
        in {
            CrosswalkStatus.COVERED,
            CrosswalkStatus.PARTIAL,
            CrosswalkStatus.CONFLICT,
        }
        and not finding.evidence
    ):
        raise HTTPException(
            status_code=422,
            detail="Evidence is required to verify this finding status.",
        )

    for field, value in patch.model_dump(exclude_unset=True).items():
        if field not in {"reviewer", "human_verified", "expected_updated_at"}:
            setattr(finding, field, value)
    if "reviewer" in patch.model_fields_set:
        finding.reviewer = patch.reviewer
    if patch.human_verified is True:
        finding.human_verified = True
        finding.reviewer = patch.reviewer
        finding.reviewed_at = utc_now()
        finding.stale = False
    elif patch.human_verified is False:
        finding.human_verified = False
        finding.reviewed_at = None
        finding.stale = False
    session.commit()
    return _finding_response(_finding(session, project_id, finding_id))


@router.post(
    "/crosswalk/{finding_id}/evidence",
    response_model=ProposalEvidenceResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_manual_evidence(
    project_id: str,
    finding_id: str,
    payload: ProposalEvidenceCreate,
    session: Session = Depends(get_session),
) -> ProposalEvidence:
    _project(session, project_id)
    finding = _finding(session, project_id, finding_id)
    document = _document(session, project_id, payload.document_id)
    if document.classification != DocumentClassification.PROPOSAL_VOLUME:
        raise HTTPException(
            status_code=422,
            detail="Crosswalk evidence must cite a proposal-volume document.",
        )
    if payload.source_end > len(document.extracted_text):
        raise HTTPException(status_code=422, detail="Evidence range exceeds the document text.")
    excerpt = document.extracted_text[payload.source_start : payload.source_end]
    if not excerpt.strip():
        raise HTTPException(status_code=422, detail="Evidence range cannot be blank.")
    evidence = ProposalEvidence(
        project_id=project_id,
        finding_id=finding.id,
        document_id=document.id,
        source_start=payload.source_start,
        source_end=payload.source_end,
        source_locator=f"characters {payload.source_start}-{payload.source_end}",
        excerpt=excerpt,
        score=1.0,
        is_manual=True,
    )
    session.add(evidence)
    if finding.human_verified:
        finding.stale = True
    finding.updated_at = utc_now()
    session.commit()
    session.refresh(evidence)
    evidence.document = document
    return evidence


@router.delete(
    "/crosswalk/{finding_id}/evidence/{evidence_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    response_class=Response,
)
def delete_manual_evidence(
    project_id: str,
    finding_id: str,
    evidence_id: str,
    session: Session = Depends(get_session),
) -> Response:
    _project(session, project_id)
    finding = _finding(session, project_id, finding_id)
    evidence = session.scalar(
        select(ProposalEvidence).where(
            ProposalEvidence.id == evidence_id,
            ProposalEvidence.finding_id == finding.id,
            ProposalEvidence.project_id == project_id,
        )
    )
    if evidence is None:
        raise HTTPException(status_code=404, detail="Proposal evidence not found.")
    if not evidence.is_manual:
        raise HTTPException(
            status_code=409,
            detail="Automated candidate evidence cannot be deleted; regenerate the crosswalk.",
        )
    session.delete(evidence)
    if finding.human_verified:
        finding.stale = True
    finding.updated_at = utc_now()
    session.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


def _stage(
    stage: WorkflowStage,
    label: str,
    completed: int,
    total: int,
    blockers: list[str],
    next_action: str | None,
) -> StageProgressResponse:
    if blockers:
        stage_status = WorkflowStatus.BLOCKED
    elif total > 0 and completed >= total:
        stage_status = WorkflowStatus.COMPLETE
    elif completed > 0:
        stage_status = WorkflowStatus.IN_PROGRESS
    else:
        stage_status = WorkflowStatus.NOT_STARTED
    return StageProgressResponse(
        stage=stage,
        label=label,
        status=stage_status,
        completed_items=completed,
        total_items=total,
        blocking_reasons=blockers,
        next_action=next_action,
    )


def _coverage_percent(compliant: int, total: int) -> float:
    if total <= 0:
        return 0.0
    if compliant >= total:
        return 100.0
    return min(round((compliant / total) * 100, 2), 99.99)


def _readiness(session: Session, project: Project) -> ReadinessResponse:
    documents = list(
        session.scalars(
            select(Document)
            .options(selectinload(Document.workflow_profile))
            .where(Document.project_id == project.id)
        )
    )
    documents_total = len(documents)
    documents_classified = sum(
        document.classification != DocumentClassification.UNCLASSIFIED for document in documents
    )
    proposal_document_records = [
        document
        for document in documents
        if document.classification == DocumentClassification.PROPOSAL_VOLUME
    ]
    proposal_documents = len(proposal_document_records)
    extraction_states = {
        state.document_id: state
        for state in session.scalars(
            select(RequirementExtractionState).where(
                RequirementExtractionState.project_id == project.id
            )
        )
    }
    classified_solicitation_documents = [
        document
        for document in documents
        if document.classification in SOLICITATION_DOCUMENT_CLASSIFICATIONS
        and document.status != DocumentStatus.ARCHIVE_EXPANDED
    ]
    solicitation_documents = len(classified_solicitation_documents)
    unusable_solicitation_documents = [
        document
        for document in classified_solicitation_documents
        if document.status != DocumentStatus.EXTRACTED
    ]
    stale_solicitation_documents = []
    for document in classified_solicitation_documents:
        if document.status != DocumentStatus.EXTRACTED:
            continue
        state = extraction_states.get(document.id)
        text_sha256 = hashlib.sha256(document.extracted_text.encode("utf-8")).hexdigest()
        if (
            state is None
            or state.blob_sha256 != document.blob_sha256
            or state.text_sha256 != text_sha256
            or state.classification != document.classification
            or state.rule_version != RULE_VERSION
        ):
            stale_solicitation_documents.append(document)

    intake = list(
        session.scalars(
            select(IntakeVerification).where(IntakeVerification.project_id == project.id)
        )
    )
    intake_verified = sum(
        item.status in {IntakeVerificationStatus.VERIFIED, IntakeVerificationStatus.NOT_APPLICABLE}
        for item in intake
    )
    intake_issues = sum(item.status == IntakeVerificationStatus.ISSUE for item in intake)

    requirements = list(
        session.scalars(
            select(Requirement)
            .join(Document, Document.id == Requirement.document_id)
            .outerjoin(DocumentProfile, DocumentProfile.document_id == Document.id)
            .where(
                Requirement.project_id == project.id,
                Requirement.validation_status != ValidationStatus.DISMISSED,
                or_(
                    DocumentProfile.document_id.is_(None),
                    DocumentProfile.classification.in_(SOLICITATION_DOCUMENT_CLASSIFICATIONS),
                ),
            )
        )
    )
    requirement_ids = {requirement.id for requirement in requirements}
    requirements_validated = sum(
        requirement.validation_status == ValidationStatus.VALIDATED for requirement in requirements
    )
    requirements_pending = len(requirements) - requirements_validated

    project_cdrls = list(
        session.scalars(
            select(CDRL)
            .options(
                selectinload(CDRL.document).selectinload(Document.workflow_profile),
                selectinload(CDRL.adjudication),
            )
            .where(CDRL.project_id == project.id)
        )
    )
    active_cdrl_adjudications = [
        _cdrl_adjudication_response(cdrl)
        for cdrl in project_cdrls
        if cdrl.document.classification != DocumentClassification.REFERENCE
    ]
    cdrls_ready = sum(item.effective_ready for item in active_cdrl_adjudications)
    cdrls_incomplete = sum(item.incomplete for item in active_cdrl_adjudications)
    cdrls_waived = sum(
        item.effective_ready and item.status == CDRLAdjudicationStatus.WAIVED
        for item in active_cdrl_adjudications
    )
    cdrls_unreviewed = sum(
        item.status == CDRLAdjudicationStatus.PENDING for item in active_cdrl_adjudications
    )
    cdrls_stale = sum(
        item.status != CDRLAdjudicationStatus.PENDING and not item.fresh
        for item in active_cdrl_adjudications
    )

    findings = list(
        session.scalars(
            select(CrosswalkFinding)
            .options(
                selectinload(CrosswalkFinding.evidence)
                .selectinload(ProposalEvidence.document)
                .selectinload(Document.workflow_profile)
            )
            .where(
                CrosswalkFinding.project_id == project.id,
                CrosswalkFinding.requirement_id.in_(requirement_ids)
                if requirement_ids
                else CrosswalkFinding.id == "",
            )
        )
    )
    invalid_evidence_finding_ids = {
        finding.id
        for finding in findings
        if (
            any(
                evidence.document.classification != DocumentClassification.PROPOSAL_VOLUME
                for evidence in finding.evidence
            )
            or (
                finding.status
                in {
                    CrosswalkStatus.COVERED,
                    CrosswalkStatus.PARTIAL,
                    CrosswalkStatus.CONFLICT,
                }
                and not finding.evidence
            )
        )
    }
    unreviewed_override_finding_ids = {
        finding.id
        for finding in findings
        if finding.status != finding.candidate_status and not finding.human_verified
    }
    stale_finding_ids = {finding.id for finding in findings if finding.stale}
    effective_findings = [
        finding
        for finding in findings
        if not finding.stale
        and finding.id not in invalid_evidence_finding_ids
        and finding.id not in unreviewed_override_finding_ids
    ]
    status_counts = {
        crosswalk_status: sum(
            finding.status == crosswalk_status for finding in effective_findings
        )
        for crosswalk_status in CrosswalkStatus
    }
    verified_findings = [
        finding
        for finding in effective_findings
        if finding.human_verified
    ]
    compliant_findings = sum(
        finding.status in {CrosswalkStatus.COVERED, CrosswalkStatus.N_A}
        for finding in effective_findings
    )
    run_state = session.get(CrosswalkRunState, project.id)
    crosswalk_inputs_fresh = bool(
        run_state is not None
        and run_state.requirement_signature == _requirement_input_signature(requirements)
        and run_state.proposal_signature == _proposal_input_signature(proposal_document_records)
    )

    actions = list(
        session.scalars(select(ProjectAction).where(ProjectAction.project_id == project.id))
    )
    actions_open = sum(action.status != ActionStatus.DONE for action in actions)
    actions_blocked = sum(action.status == ActionStatus.BLOCKED for action in actions)

    file_blockers: list[str] = []
    if solicitation_documents == 0:
        file_blockers.append("Upload at least one solicitation document.")
    if documents_classified < documents_total:
        file_blockers.append("Classify every uploaded document.")
    requirement_blockers: list[str] = []
    if unusable_solicitation_documents:
        requirement_blockers.append(
            "Resolve "
            f"{len(unusable_solicitation_documents)} solicitation document extraction issue(s)."
        )
    if stale_solicitation_documents:
        requirement_blockers.append(
            "Run requirement extraction for "
            f"{len(stale_solicitation_documents)} new or reclassified solicitation document(s)."
        )
    if not requirements:
        requirement_blockers.append("Extract requirement candidates from the solicitation.")
    unusable_proposal_documents = sum(
        document.status not in {DocumentStatus.EXTRACTED, DocumentStatus.ARCHIVE_EXPANDED}
        for document in proposal_document_records
    )
    crosswalk_blockers: list[str] = []
    if not proposal_documents:
        crosswalk_blockers.append("Upload and classify at least one proposal volume.")
    if unusable_proposal_documents:
        crosswalk_blockers.append(
            f"Resolve {unusable_proposal_documents} proposal document extraction issue(s)."
        )
    if len(findings) < len(requirements):
        crosswalk_blockers.append("Generate the proposal crosswalk.")
    elif requirements and not crosswalk_inputs_fresh:
        crosswalk_blockers.append(
            "Regenerate the proposal crosswalk after requirement or proposal changes."
        )
    if invalid_evidence_finding_ids:
        crosswalk_blockers.append(
            "Replace invalid proposal evidence on "
            f"{len(invalid_evidence_finding_ids)} crosswalk finding(s)."
        )
    if unreviewed_override_finding_ids:
        crosswalk_blockers.append(
            "Human-review "
            f"{len(unreviewed_override_finding_ids)} manually overridden crosswalk finding(s)."
        )
    if stale_finding_ids and crosswalk_inputs_fresh:
        crosswalk_blockers.append(
            f"Review {len(stale_finding_ids)} stale crosswalk finding(s) after input changes."
        )
    unverified = sum(
        not finding.human_verified
        or finding.stale
        or finding.id in invalid_evidence_finding_ids
        or finding.id in unreviewed_override_finding_ids
        for finding in findings
    )
    gap_count = sum(
        finding.status
        in {CrosswalkStatus.PARTIAL, CrosswalkStatus.MISSING, CrosswalkStatus.CONFLICT}
        for finding in effective_findings
    )
    if gap_count:
        crosswalk_blockers.append(f"Resolve {gap_count} proposal coverage gap(s).")

    stages = [
        _stage(
            WorkflowStage.SOLICITATION_FILES,
            "Solicitation files",
            documents_classified,
            max(documents_total, 1),
            file_blockers,
            "Upload and classify the solicitation package." if file_blockers else None,
        ),
        _stage(
            WorkflowStage.REQUIREMENTS,
            "Requirements",
            len(requirements),
            max(len(requirements), 1),
            requirement_blockers,
            "Extract the current solicitation requirements." if requirement_blockers else None,
        ),
        _stage(
            WorkflowStage.CROSSWALK,
            "Crosswalk",
            compliant_findings,
            max(len(requirements), 1),
            crosswalk_blockers,
            "Generate or review the proposal crosswalk." if crosswalk_blockers else None,
        ),
    ]
    blockers = [reason for stage in stages for reason in stage.blocking_reasons]
    # This percentage is proposal coverage, not workflow progress. Source-file and
    # extraction steps must not inflate the result shown to proposal reviewers.
    readiness_percent = _coverage_percent(compliant_findings, len(requirements))
    ready = (
        not blockers
        and bool(requirements)
        and len(effective_findings) == len(requirements)
        and compliant_findings == len(requirements)
    )
    authoritative_stage = next(
        (stage for stage in stages if stage.status != WorkflowStatus.COMPLETE),
        stages[-1],
    )
    next_action = next(
        (stage.next_action for stage in stages if stage.blocking_reasons and stage.next_action),
        None,
    )
    return ReadinessResponse(
        project_id=project.id,
        ready=ready,
        readiness_percent=readiness_percent,
        workflow_stage=authoritative_stage.stage,
        workflow_status=authoritative_stage.status,
        documents_total=documents_total,
        documents_classified=documents_classified,
        proposal_documents=proposal_documents,
        intake_total=len(intake),
        intake_verified=intake_verified,
        intake_issues=intake_issues,
        requirements_total=len(requirements),
        requirements_validated=requirements_validated,
        requirements_pending=requirements_pending,
        cdrls_total=len(active_cdrl_adjudications),
        cdrls_ready=cdrls_ready,
        cdrls_incomplete=cdrls_incomplete,
        cdrls_unreviewed=cdrls_unreviewed,
        cdrls_waived=cdrls_waived,
        cdrls_stale=cdrls_stale,
        crosswalk_total=len(effective_findings),
        crosswalk_verified=len(verified_findings),
        covered=status_counts[CrosswalkStatus.COVERED],
        partial=status_counts[CrosswalkStatus.PARTIAL],
        missing=status_counts[CrosswalkStatus.MISSING],
        conflict=status_counts[CrosswalkStatus.CONFLICT],
        n_a=status_counts[CrosswalkStatus.N_A],
        unverified=unverified,
        actions_open=actions_open,
        actions_blocked=actions_blocked,
        blocking_reasons=blockers,
        next_action=next_action,
        stages=stages,
    )


@router.get("/readiness", response_model=ReadinessResponse)
def get_readiness(project_id: str, session: Session = Depends(get_session)) -> ReadinessResponse:
    return _readiness(session, _project(session, project_id))


def _requirement_rows(
    session: Session, project_id: str, section: RequirementSection | None = None
) -> list[dict[str, Any]]:
    statement = (
        select(Requirement)
        .join(Document, Document.id == Requirement.document_id)
        .outerjoin(DocumentProfile, DocumentProfile.document_id == Document.id)
        .options(selectinload(Requirement.document))
        .where(
            Requirement.project_id == project_id,
            or_(
                DocumentProfile.document_id.is_(None),
                DocumentProfile.classification.in_(SOLICITATION_DOCUMENT_CLASSIFICATIONS),
            ),
        )
    )
    if section is not None:
        statement = statement.where(Requirement.section == section)
    requirements = session.scalars(
        statement.order_by(Requirement.section, Requirement.document_id, Requirement.source_start)
    )
    return [
        {
            "id": item.id,
            "section": item.section.value,
            "category": item.category.value,
            "requirement": item.requirement_text,
            "source_document": item.document.name,
            "source_locator": item.source_locator,
            "source_excerpt": item.source_text,
            "owner": item.obligation_owner.value,
            "applicability": item.applicability.value,
            "validation_status": item.validation_status.value,
            "reviewer": item.reviewer,
            "review_note": item.review_note,
        }
        for item in requirements
    ]


def _cdrl_rows(session: Session, project_id: str) -> list[dict[str, Any]]:
    cdrls = session.scalars(
        select(CDRL)
        .join(Document, Document.id == CDRL.document_id)
        .outerjoin(DocumentProfile, DocumentProfile.document_id == Document.id)
        .options(
            selectinload(CDRL.document).selectinload(Document.workflow_profile),
            selectinload(CDRL.requirement),
            selectinload(CDRL.adjudication),
        )
        .where(
            CDRL.project_id == project_id,
            or_(
                DocumentProfile.document_id.is_(None),
                DocumentProfile.classification.in_(SOLICITATION_DOCUMENT_CLASSIFICATIONS),
            ),
        )
        .order_by(CDRL.document_id, CDRL.source_start)
    )
    rows = []
    for item in cdrls:
        row = cdrl_response(item).model_dump(mode="json")
        row.update(_cdrl_adjudication_response(item).model_dump(mode="json"))
        rows.append(row)
    return rows


def _crosswalk_rows(session: Session, project_id: str) -> list[dict[str, Any]]:
    findings = list_crosswalk(
        project_id,
        finding_status=None,
        human_verified=None,
        session=session,
    )
    rows: list[dict[str, Any]] = []
    for finding in findings:
        evidence = [
            {
                "document": item.document.name,
                "locator": item.source_locator,
                "excerpt": item.excerpt,
                "score": item.score,
                "manual": item.is_manual,
            }
            for item in finding.evidence
        ]
        rows.append(
            {
                "id": finding.id,
                "requirement_id": finding.requirement_id,
                "section": finding.requirement.section.value,
                "requirement": finding.requirement.requirement_text,
                "candidate_status": finding.candidate_status.value,
                "status": finding.status.value,
                "score": finding.score,
                "human_verified": finding.human_verified,
                "stale": finding.stale,
                "reviewer": finding.reviewer,
                "owner": finding.owner,
                "due_at": finding.due_at.isoformat() if finding.due_at else None,
                "notes": finding.notes,
                "evidence": evidence,
            }
        )
    return rows


def _export_data(session: Session, project: Project, register_name: str) -> list[dict[str, Any]]:
    normalized = register_name.lower()
    if normalized == "requirements":
        return _requirement_rows(session, project.id)
    if normalized in {"section-l", "l"}:
        return _requirement_rows(session, project.id, RequirementSection.L)
    if normalized in {"section-m", "m"}:
        return _requirement_rows(session, project.id, RequirementSection.M)
    if normalized == "cdrls":
        return _cdrl_rows(session, project.id)
    if normalized == "crosswalk":
        return _crosswalk_rows(session, project.id)
    if normalized == "readiness":
        return [_readiness(session, project).model_dump(mode="json")]
    raise HTTPException(status_code=404, detail="Export register not found.")


def _flat_value(value: Any) -> str | int | float | bool:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        result: str | int | float | bool = value
    else:
        result = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    if isinstance(result, str) and result.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + result
    return result


def _csv_bytes(rows: list[dict[str, Any]]) -> bytes:
    stream = io.StringIO(newline="")
    if not rows:
        stream.write("\ufeff")
        return stream.getvalue().encode("utf-8")
    fieldnames = list(dict.fromkeys(key for row in rows for key in row))
    writer = csv.DictWriter(stream, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({key: _flat_value(value) for key, value in row.items()})
    return ("\ufeff" + stream.getvalue()).encode("utf-8")


@router.get("/exports/workbook.xlsx")
def export_workbook(project_id: str, session: Session = Depends(get_session)) -> StreamingResponse:
    project = _project(session, project_id)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    sheet_specs = (
        ("Requirements", _requirement_rows(session, project.id)),
        ("Section L", _requirement_rows(session, project.id, RequirementSection.L)),
        ("Section M", _requirement_rows(session, project.id, RequirementSection.M)),
        ("CDRLs", _cdrl_rows(session, project.id)),
        ("Crosswalk", _crosswalk_rows(session, project.id)),
        ("Readiness", [_readiness(session, project).model_dump(mode="json")]),
    )
    for title, rows in sheet_specs:
        sheet = workbook.create_sheet(title)
        fieldnames = list(dict.fromkeys(key for row in rows for key in row))
        if not fieldnames:
            fieldnames = ["No records"]
        sheet.append(fieldnames)
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill("solid", fgColor="17365D")
        for row in rows:
            sheet.append([_flat_value(row.get(field)) for field in fieldnames])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
            sheet.column_dimensions[column[0].column_letter].width = max(width, 12)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="compliance-{project.id[:8]}-workbook.xlsx"'
            )
        },
    )


@router.get("/exports/{register_name}", response_model=None)
def export_register(
    project_id: str,
    register_name: str,
    export_format: str = Query(default="json", alias="format", pattern="^(json|csv)$"),
    session: Session = Depends(get_session),
) -> JSONResponse | StreamingResponse:
    project = _project(session, project_id)
    rows = _export_data(session, project, register_name)
    filename = f"compliance-{register_name}.{export_format}"
    if export_format == "csv":
        return StreamingResponse(
            io.BytesIO(_csv_bytes(rows)),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    return JSONResponse(
        content=jsonable_encoder(rows),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
