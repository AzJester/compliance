from __future__ import annotations

import io
import sqlite3
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from threading import Event

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from backend.app import workflow_api
from backend.app.config import Settings
from backend.app.database import create_database, initialize_database
from backend.app.main import create_app
from backend.app.models import CDRL, Blob, Document, DocumentStatus, Project
from backend.app.workflow_api import _coverage_percent

from .conftest import make_test_client, pdf_bytes, seed_extracted_document


def _create_project(client: TestClient, name: str = "Synthetic RFP") -> dict[str, object]:
    response = client.post(
        "/api/projects",
        json={
            "name": name,
            "solicitation_number": "FAKE-26-R-0001",
            "agency": "Synthetic Agency",
            "due_at": "2026-10-01T17:00:00Z",
            "due_timezone": "UTC",
        },
    )
    assert response.status_code == 201, response.text
    return response.json()


def test_coverage_percent_never_rounds_incomplete_coverage_to_100() -> None:
    assert _coverage_percent(1_249, 1_250) == 99.92
    assert _coverage_percent(19_999, 20_000) == 99.99
    assert _coverage_percent(1_250, 1_250) == 100.0


def _classify(
    client: TestClient,
    project_id: str,
    document_id: str,
    classification: str,
    *,
    volume_name: str | None = None,
) -> dict[str, object]:
    response = client.patch(
        f"/api/projects/{project_id}/documents/{document_id}/profile",
        json={"classification": classification, "volume_name": volume_name},
    )
    assert response.status_code == 200, response.text
    return response.json()


def _extract(client: TestClient, project_id: str) -> dict[str, object]:
    response = client.post(f"/api/projects/{project_id}/requirements/extract")
    assert response.status_code == 200, response.text
    return response.json()


def _pending_crosswalk_project(client: TestClient, *, name: str, requirement_text: str) -> str:
    project_id = str(_create_project(client, name)["id"])
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        f"SECTION C\n{requirement_text}",
        name=f"{name}-solicitation.pdf",
    )
    proposal_id = seed_extracted_document(
        client,
        project_id,
        requirement_text,
        name=f"{name}-proposal.pdf",
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    _classify(
        client,
        project_id,
        proposal_id,
        "PROPOSAL_VOLUME",
        volume_name="Technical",
    )
    summary = _extract(client, project_id)
    assert summary["total_requirements"] == 1
    return project_id


def _cdrl_text(*, complete: bool) -> str:
    values = {
        "A": "0001",
        "B": "A",
        "C": "TDP",
        "D": "AIR VEHICLE",
        "E": "FAKE-26-C-0001",
        "F": "SYNTHETIC CONTRACTOR",
        "1": "A001",
        "2": "MONTHLY STATUS REPORT",
        "3": "TECHNICAL PROGRESS",
        "4": "DI-MGMT-80368",
        "5": "PWS 3.2.1",
        "6": "SYNTHETIC/OFFICE",
        "7": "LT",
        "8": "A",
        "9": "D",
        "10": "MTHLY",
        "11": "20261001",
        "12": "030 DAC",
        "13": "15 DARP",
        "14": "Draft 01; Final 02",
        "15": "003",
        "16": "Synthetic tailoring only.",
        "17": "A",
        "18": "$000000",
    }
    if not complete:
        values = {"1": values["1"], "2": values["2"]}
    return "DD FORM 1423\n" + "\n".join(
        f"Block {block}: {value}" for block, value in values.items()
    )


def _extract_one_cdrl(
    client: TestClient, project_id: str, *, complete: bool
) -> tuple[str, dict[str, object]]:
    document_id = seed_extracted_document(
        client,
        project_id,
        _cdrl_text(complete=complete),
        name="synthetic-cdrl.pdf",
    )
    _classify(client, project_id, document_id, "CDRL")
    _extract(client, project_id)
    cdrls = client.get(f"/api/projects/{project_id}/cdrls").json()
    assert len(cdrls) == 1
    return document_id, cdrls[0]


def _ready_project(
    client: TestClient,
    *,
    include_backup_proposal: bool = False,
) -> dict[str, object]:
    project = _create_project(client, "Freshness Test")
    project_id = str(project["id"])
    requirement_text = "The contractor shall deliver ten monthly status reports."
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        f"SECTION C\n{requirement_text}",
        name="solicitation.pdf",
    )
    proposal_id = seed_extracted_document(
        client,
        project_id,
        requirement_text,
        name="proposal-primary.pdf",
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    _classify(
        client,
        project_id,
        proposal_id,
        "PROPOSAL_VOLUME",
        volume_name="Technical",
    )
    backup_proposal_id = None
    if include_backup_proposal:
        backup_proposal_id = seed_extracted_document(
            client,
            project_id,
            requirement_text,
            name="proposal-backup.pdf",
        )
        _classify(
            client,
            project_id,
            backup_proposal_id,
            "PROPOSAL_VOLUME",
            volume_name="Management",
        )

    _extract(client, project_id)
    requirement = client.get(f"/api/projects/{project_id}/requirements").json()[0]
    response = client.patch(
        f"/api/projects/{project_id}/requirements/{requirement['id']}",
        json={
            "validation_status": "VALIDATED",
            "reviewer": "Requirements Reviewer",
            "expected_updated_at": requirement["updated_at"],
        },
    )
    assert response.status_code == 200, response.text
    requirement = response.json()

    checks = client.post(f"/api/projects/{project_id}/intake-verifications/initialize").json()
    for check in checks:
        response = client.patch(
            f"/api/projects/{project_id}/intake-verifications/{check['id']}",
            json={"status": "VERIFIED", "reviewer": "Intake Reviewer"},
        )
        assert response.status_code == 200, response.text

    response = client.post(f"/api/projects/{project_id}/crosswalk/generate")
    assert response.status_code == 200, response.text
    finding = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    response = client.patch(
        f"/api/projects/{project_id}/crosswalk/{finding['id']}",
        json={"status": "COVERED", "human_verified": True, "reviewer": "Reviewer"},
    )
    assert response.status_code == 200, response.text
    assert client.get(f"/api/projects/{project_id}/readiness").json()["ready"] is True
    return {
        "project_id": project_id,
        "solicitation_id": solicitation_id,
        "proposal_id": proposal_id,
        "backup_proposal_id": backup_proposal_id,
        "requirement": requirement,
        "finding": response.json(),
    }


def test_project_workflow_metadata_document_profile_and_bounded_text(
    client: TestClient,
) -> None:
    project = _create_project(client)
    project_id = str(project["id"])
    document_id = seed_extracted_document(
        client,
        project_id,
        "0123456789abcdefghij",
        name="public-synthetic.pdf",
    )

    updated = client.patch(
        f"/api/projects/{project_id}",
        json={"name": "Updated Synthetic RFP", "agency": None},
    )
    assert updated.status_code == 200
    assert updated.json()["name"] == "Updated Synthetic RFP"
    assert updated.json()["agency"] is None
    assert client.patch(f"/api/projects/{project_id}", json={}).status_code == 422
    assert (
        client.patch(
            f"/api/projects/{project_id}",
            json={"due_timezone": None},
        ).status_code
        == 422
    )

    workflow = client.get(f"/api/projects/{project_id}/workflow")
    assert workflow.status_code == 200
    assert workflow.json()["stage"] == "PROJECT_SETUP"
    assert workflow.json()["status"] == "IN_PROGRESS"
    assert (
        client.patch(
            f"/api/projects/{project_id}/workflow",
            json={"status": "BLOCKED"},
        ).status_code
        == 422
    )
    blocked = client.patch(
        f"/api/projects/{project_id}/workflow",
        json={"stage": "VERIFY_PACKAGE", "status": "BLOCKED", "blocker_summary": "Test"},
    )
    assert blocked.status_code == 200
    assert blocked.json()["blocker_summary"] == "Test"

    unclassified = client.get(f"/api/projects/{project_id}/documents").json()[0]
    assert unclassified["classification"] == "UNCLASSIFIED"
    assert unclassified["volume_name"] is None
    assert (
        client.patch(
            f"/api/projects/{project_id}/documents/{document_id}/profile",
            json={"classification": "PROPOSAL_VOLUME"},
        ).status_code
        == 422
    )
    profile = _classify(client, project_id, document_id, "ATTACHMENT")
    assert profile["classification"] == "ATTACHMENT"
    assert "storage_path" not in profile

    text = client.get(f"/api/projects/{project_id}/documents/{document_id}/text?start=5&limit=7")
    assert text.status_code == 200
    assert text.json() == {
        "document_id": document_id,
        "name": "public-synthetic.pdf",
        "total_characters": 20,
        "start": 5,
        "end": 12,
        "text": "56789ab",
        "truncated": True,
    }

    other = _create_project(client, "Other")
    assert (
        client.get(f"/api/projects/{other['id']}/documents/{document_id}/text").status_code == 404
    )


def test_intake_checklist_and_actions_are_idempotent_and_project_scoped(
    client: TestClient,
) -> None:
    project = _create_project(client)
    project_id = str(project["id"])
    initialized = client.post(f"/api/projects/{project_id}/intake-verifications/initialize")
    assert initialized.status_code == 200
    checks = initialized.json()
    assert len(checks) == 6
    assert (
        len(client.post(f"/api/projects/{project_id}/intake-verifications/initialize").json()) == 6
    )

    verified = client.patch(
        f"/api/projects/{project_id}/intake-verifications/{checks[0]['id']}",
        json={"status": "VERIFIED", "reviewer": "Synthetic Reviewer", "note": "Checked"},
    )
    assert verified.status_code == 200
    assert verified.json()["status"] == "VERIFIED"
    duplicate = client.post(
        f"/api/projects/{project_id}/intake-verifications",
        json={"check_key": checks[0]["check_key"], "label": "Duplicate"},
    )
    assert duplicate.status_code == 409

    action = client.post(
        f"/api/projects/{project_id}/actions",
        json={
            "title": "Resolve package issue",
            "description": "Synthetic task",
            "owner": "Proposal Lead",
            "due_at": "2026-09-20T12:00:00Z",
        },
    )
    assert action.status_code == 201
    action_id = action.json()["id"]
    patched = client.patch(
        f"/api/projects/{project_id}/actions/{action_id}",
        json={"status": "BLOCKED"},
    )
    assert patched.status_code == 200
    assert patched.json()["status"] == "BLOCKED"
    assert client.get(f"/api/projects/{project_id}/actions").json()[0]["id"] == action_id

    other = _create_project(client, "Other")
    assert (
        client.patch(
            f"/api/projects/{other['id']}/actions/{action_id}",
            json={"status": "DONE"},
        ).status_code
        == 404
    )


def test_proposal_documents_are_excluded_from_requirement_extraction_and_crosswalked(
    client: TestClient,
) -> None:
    project = _create_project(client)
    project_id = str(project["id"])
    requirement_text = "The contractor shall deliver ten monthly status reports."
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        f"SECTION C - PERFORMANCE WORK STATEMENT\n{requirement_text}",
        name="solicitation.pdf",
    )
    proposal_id = seed_extracted_document(
        client,
        project_id,
        requirement_text,
        name="technical-volume.pdf",
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    _classify(
        client,
        project_id,
        proposal_id,
        "PROPOSAL_VOLUME",
        volume_name="Volume I - Technical",
    )

    summary = _extract(client, project_id)
    assert summary["documents_analyzed"] == 1
    requirements = client.get(f"/api/projects/{project_id}/requirements").json()
    assert len(requirements) == 1
    assert requirements[0]["document_id"] == solicitation_id

    generated = client.post(f"/api/projects/{project_id}/crosswalk/generate")
    assert generated.status_code == 200
    assert generated.json() == {
        "requirements_analyzed": 1,
        "proposal_documents_analyzed": 1,
        "findings_created": 1,
        "findings_updated": 0,
        "verified_findings_marked_stale": 0,
    }
    finding = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    assert finding["candidate_status"] == "COVERED"
    assert finding["status"] == "COVERED"
    assert finding["human_verified"] is False
    assert finding["needs_attention"] is False
    assert finding["attention_reasons"] == []
    assert finding["evidence"][0]["document_id"] == proposal_id
    assert finding["evidence"][0]["excerpt"] == requirement_text
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["ready"] is True
    assert readiness["readiness_percent"] == 100
    assert readiness["requirements_pending"] == 1
    assert readiness["crosswalk_verified"] == 0
    assert readiness["unverified"] == 1
    assert readiness["blocking_reasons"] == []
    assert [stage["stage"] for stage in readiness["stages"]] == [
        "SOLICITATION_FILES",
        "REQUIREMENTS",
        "CROSSWALK",
    ]

    assert (
        client.patch(
            f"/api/projects/{project_id}/crosswalk/{finding['id']}",
            json={"human_verified": True},
        ).status_code
        == 422
    )
    verified = client.patch(
        f"/api/projects/{project_id}/crosswalk/{finding['id']}",
        json={
            "status": "COVERED",
            "human_verified": True,
            "reviewer": "Human Reviewer",
            "owner": "Volume Lead",
            "notes": "Exact synthetic match.",
        },
    )
    assert verified.status_code == 200
    assert verified.json()["human_verified"] is True
    assert verified.json()["reviewed_at"] is not None
    stale_update = client.patch(
        f"/api/projects/{project_id}/crosswalk/{finding['id']}",
        json={
            "notes": "This stale edit must not persist.",
            "expected_updated_at": finding["updated_at"],
        },
    )
    assert stale_update.status_code == 409
    assert "refresh" in stale_update.text.lower()
    assert client.get(f"/api/projects/{project_id}/readiness").json()["crosswalk_verified"] == 1


def test_crosswalk_generation_is_project_scoped_and_rejects_duplicate_work(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    first_project_id = _pending_crosswalk_project(
        client,
        name="Concurrent First",
        requirement_text="The contractor shall deliver the alpha management plan.",
    )
    other_project_id = _pending_crosswalk_project(
        client,
        name="Concurrent Other",
        requirement_text="The contractor shall deliver the bravo staffing plan.",
    )
    started = Event()
    release = Event()
    original_proposal_corpus = workflow_api._proposal_corpus

    def blocking_proposal_corpus(
        documents: list[workflow_api.Document],
    ) -> workflow_api._ProposalCorpus:
        if documents and documents[0].project_id == first_project_id:
            started.set()
            assert release.wait(timeout=10)
        return original_proposal_corpus(documents)

    monkeypatch.setattr(workflow_api, "_proposal_corpus", blocking_proposal_corpus)
    first_url = f"/api/projects/{first_project_id}/crosswalk/generate"
    with ThreadPoolExecutor(max_workers=1) as executor:
        first_request = executor.submit(client.post, first_url)
        assert started.wait(timeout=10)
        try:
            duplicate = client.post(first_url)
            other_project = client.post(f"/api/projects/{other_project_id}/crosswalk/generate")
        finally:
            release.set()
        first = first_request.result(timeout=10)

    assert duplicate.status_code == 409
    assert "already running" in duplicate.json()["detail"]
    assert other_project.status_code == 200, other_project.text
    assert first.status_code == 200, first.text
    assert len(client.get(f"/api/projects/{first_project_id}/crosswalk").json()) == 1


def test_crosswalk_generation_guard_is_released_after_failure(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    project_id = _pending_crosswalk_project(
        client,
        name="Retry After Failure",
        requirement_text="The contractor shall deliver the quality plan.",
    )
    original_proposal_corpus = workflow_api._proposal_corpus
    failed = False

    def fail_once(
        documents: list[workflow_api.Document],
    ) -> workflow_api._ProposalCorpus:
        nonlocal failed
        if not failed:
            failed = True
            raise RuntimeError("injected proposal analysis failure")
        return original_proposal_corpus(documents)

    monkeypatch.setattr(workflow_api, "_proposal_corpus", fail_once)
    url = f"/api/projects/{project_id}/crosswalk/generate"

    with pytest.raises(RuntimeError, match="injected proposal analysis failure"):
        client.post(url)

    retry = client.post(url)
    assert retry.status_code == 200, retry.text


def test_sqlite_lock_during_crosswalk_commit_returns_retryable_conflict(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    project_id = _pending_crosswalk_project(
        client,
        name="SQLite Lock",
        requirement_text="The contractor shall deliver the security plan.",
    )
    original_commit = workflow_api.Session.commit
    lock_injected = False

    def locked_once(session: workflow_api.Session) -> None:
        nonlocal lock_injected
        generating_findings = any(
            isinstance(item, workflow_api.CrosswalkFinding) for item in session.new
        )
        if generating_findings and not lock_injected:
            lock_injected = True
            raise workflow_api.OperationalError(
                "INSERT INTO crosswalk_findings",
                {},
                sqlite3.OperationalError("database is locked"),
            )
        original_commit(session)

    monkeypatch.setattr(workflow_api.Session, "commit", locked_once)
    url = f"/api/projects/{project_id}/crosswalk/generate"

    locked = client.post(url)
    retry = client.post(url)

    assert locked.status_code == 409
    assert "already running" in locked.json()["detail"]
    assert retry.status_code == 200, retry.text


def test_crosswalk_generation_flushes_scale_batch_once(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    project_id = str(_create_project(client, "Batched Crosswalk")["id"])
    requirement_lines = [
        f"The contractor shall deliver synthetic report number {index} every month."
        for index in range(40)
    ]
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        "SECTION C\n" + "\n".join(requirement_lines),
        name="batched-solicitation.pdf",
    )
    proposal_id = seed_extracted_document(
        client,
        project_id,
        "\n".join(requirement_lines),
        name="batched-proposal.pdf",
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    _classify(
        client,
        project_id,
        proposal_id,
        "PROPOSAL_VOLUME",
        volume_name="Technical",
    )
    assert _extract(client, project_id)["total_requirements"] == 40
    original_flush = workflow_api.Session.flush
    write_flushes = 0

    def counting_flush(session: workflow_api.Session, objects: object | None = None) -> None:
        nonlocal write_flushes
        if session.new or session.dirty or session.deleted:
            write_flushes += 1
        original_flush(session, objects)

    monkeypatch.setattr(workflow_api.Session, "flush", counting_flush)

    generated = client.post(f"/api/projects/{project_id}/crosswalk/generate")

    assert generated.status_code == 200, generated.text
    assert generated.json()["findings_created"] == 40
    assert write_flushes == 1
    assert len(client.get(f"/api/projects/{project_id}/crosswalk").json()) == 40


def test_proposal_upload_classification_is_atomic(client: TestClient) -> None:
    project = _create_project(client)
    project_id = str(project["id"])
    response = client.post(
        f"/api/projects/{project_id}/documents",
        data={
            "classification": "PROPOSAL_VOLUME",
            "volume_name": "Volume I - Technical",
            "classification_notes": "Synthetic response only.",
        },
        files={
            "files": (
                "proposal.pdf",
                pdf_bytes("The contractor shall provide the synthetic plan."),
                "application/pdf",
            )
        },
    )
    assert response.status_code == 201, response.text
    uploaded = response.json()[0]
    assert uploaded["classification"] == "PROPOSAL_VOLUME"
    assert uploaded["volume_name"] == "Volume I - Technical"
    assert uploaded["classification_notes"] == "Synthetic response only."
    assert _extract(client, project_id)["documents_analyzed"] == 0
    missing_volume_name = client.post(
        f"/api/projects/{project_id}/documents",
        data={"classification": "PROPOSAL_VOLUME"},
        files={
            "files": (
                "invalid.pdf",
                pdf_bytes("Synthetic"),
                "application/pdf",
            )
        },
    )
    assert missing_volume_name.status_code == 422


def test_blank_extracted_proposal_is_rejected_and_ignored_when_valid_volume_exists(
    client: TestClient,
) -> None:
    project_id = str(_create_project(client, "Blank Proposal Text")["id"])
    requirement_text = "The contractor shall deliver the synthetic quality plan."
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        f"SECTION C\n{requirement_text}",
        name="blank-proposal-solicitation.pdf",
    )
    blank_proposal_id = seed_extracted_document(
        client,
        project_id,
        "",
        name="blank-proposal.docx",
        content_type=("application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    _classify(
        client,
        project_id,
        blank_proposal_id,
        "PROPOSAL_VOLUME",
        volume_name="Blank legacy volume",
    )
    assert _extract(client, project_id)["total_requirements"] == 1

    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    rejected = client.post(f"/api/projects/{project_id}/crosswalk/generate")

    assert any(
        "proposal document extraction issue" in reason for reason in readiness["blocking_reasons"]
    )
    assert rejected.status_code == 422
    assert "usable extracted text" in rejected.json()["detail"]

    valid_proposal_id = seed_extracted_document(
        client,
        project_id,
        requirement_text,
        name="searchable-proposal.pdf",
    )
    _classify(
        client,
        project_id,
        valid_proposal_id,
        "PROPOSAL_VOLUME",
        volume_name="Searchable technical volume",
    )
    generated = client.post(f"/api/projects/{project_id}/crosswalk/generate")
    finding = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]

    assert generated.status_code == 200, generated.text
    assert generated.json()["proposal_documents_analyzed"] == 1
    assert finding["evidence"][0]["document_id"] == valid_proposal_id


def test_crosswalk_generation_requires_a_proposal_volume(client: TestClient) -> None:
    project_id = str(_create_project(client, "No Proposal Volume")["id"])
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        "SECTION C\nThe contractor shall deliver the management plan.",
        name="no-proposal-solicitation.pdf",
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    assert _extract(client, project_id)["total_requirements"] == 1

    response = client.post(f"/api/projects/{project_id}/crosswalk/generate")

    assert response.status_code == 422
    assert "Upload and classify" in response.json()["detail"]


def test_confirmation_only_rerun_adopts_fresh_automated_result(client: TestClient) -> None:
    project = _create_project(client)
    project_id = str(project["id"])
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        "SECTION C\nThe contractor shall deliver 10 reports monthly.",
        name="solicitation.pdf",
    )
    proposal_id = seed_extracted_document(
        client,
        project_id,
        "The contractor shall deliver 5 reports monthly.",
        name="proposal.pdf",
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    _classify(client, project_id, proposal_id, "PROPOSAL_VOLUME", volume_name="Technical")
    _extract(client, project_id)
    client.post(f"/api/projects/{project_id}/crosswalk/generate")
    finding = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    assert finding["candidate_status"] == "CONFLICT"
    verified = client.patch(
        f"/api/projects/{project_id}/crosswalk/{finding['id']}",
        json={"human_verified": True, "reviewer": "Reviewer", "status": "CONFLICT"},
    )
    assert verified.status_code == 200

    with client.app.state.session_factory() as session:
        proposal = session.scalar(select(Document).where(Document.id == proposal_id))
        assert proposal is not None
        proposal.extracted_text = "The contractor shall deliver 10 reports monthly."
        proposal.extraction_count = len(proposal.extracted_text)
        session.commit()

    rerun = client.post(f"/api/projects/{project_id}/crosswalk/generate")
    assert rerun.status_code == 200
    assert rerun.json()["verified_findings_marked_stale"] == 0
    refreshed = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    assert refreshed["candidate_status"] == "COVERED"
    assert refreshed["status"] == "COVERED"
    assert refreshed["human_verified"] is False
    assert refreshed["stale"] is False
    assert refreshed["needs_attention"] is False
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["crosswalk_verified"] == 0
    assert readiness["unverified"] == 1
    assert readiness["ready"] is True
    assert readiness["blocking_reasons"] == []


def test_crosswalk_matches_evidence_wrapped_across_extracted_lines(client: TestClient) -> None:
    project = _create_project(client)
    project_id = str(project["id"])
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        "SECTION C\nThe contractor shall deliver monthly status reports.",
        name="solicitation.pdf",
    )
    proposal_id = seed_extracted_document(
        client,
        project_id,
        "The contractor shall deliver\nmonthly status reports.",
        name="proposal.pdf",
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    _classify(client, project_id, proposal_id, "PROPOSAL_VOLUME", volume_name="Technical")
    _extract(client, project_id)
    assert client.post(f"/api/projects/{project_id}/crosswalk/generate").status_code == 200
    finding = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    assert finding["candidate_status"] == "COVERED"
    assert "deliver\nmonthly" in finding["evidence"][0]["excerpt"]


def test_manual_evidence_is_exact_bounded_and_proposal_only(client: TestClient) -> None:
    project = _create_project(client)
    project_id = str(project["id"])
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        "SECTION C\nThe contractor shall provide a safety plan.",
        name="solicitation.pdf",
    )
    proposal_text = "Introduction. Our safety plan provides controls and evidence. Conclusion."
    proposal_id = seed_extracted_document(
        client,
        project_id,
        proposal_text,
        name="proposal.pdf",
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    _classify(client, project_id, proposal_id, "PROPOSAL_VOLUME", volume_name="Technical")
    _extract(client, project_id)
    client.post(f"/api/projects/{project_id}/crosswalk/generate")
    finding = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]

    assert (
        client.post(
            f"/api/projects/{project_id}/crosswalk/{finding['id']}/evidence",
            json={"document_id": solicitation_id, "source_start": 0, "source_end": 10},
        ).status_code
        == 422
    )
    start = proposal_text.index("Our safety")
    end = proposal_text.index(" Conclusion")
    evidence = client.post(
        f"/api/projects/{project_id}/crosswalk/{finding['id']}/evidence",
        json={"document_id": proposal_id, "source_start": start, "source_end": end},
    )
    assert evidence.status_code == 201
    assert evidence.json()["excerpt"] == proposal_text[start:end]
    assert evidence.json()["is_manual"] is True
    assert "storage_path" not in evidence.json()
    assert (
        client.post(
            f"/api/projects/{project_id}/crosswalk/{finding['id']}/evidence",
            json={
                "document_id": proposal_id,
                "source_start": 0,
                "source_end": len(proposal_text) + 1,
            },
        ).status_code
        == 422
    )


def test_automated_crosswalk_can_reach_ready_without_approval_clicks(
    client: TestClient,
) -> None:
    project = _create_project(client)
    project_id = str(project["id"])
    requirement_text = "The offeror shall submit a transition plan."
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        f"SECTION L - INSTRUCTIONS\n{requirement_text}",
        name="rfp.pdf",
    )
    proposal_id = seed_extracted_document(
        client,
        project_id,
        requirement_text,
        name="volume-i.pdf",
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    _classify(client, project_id, proposal_id, "PROPOSAL_VOLUME", volume_name="Volume I")
    _extract(client, project_id)
    requirement = client.get(f"/api/projects/{project_id}/requirements").json()[0]
    assert requirement["validation_status"] == "PENDING"
    assert client.get(f"/api/projects/{project_id}/intake-verifications").json() == []
    generated = client.post(f"/api/projects/{project_id}/crosswalk/generate")
    assert generated.status_code == 200
    readiness = client.get(f"/api/projects/{project_id}/readiness")
    assert readiness.status_code == 200
    assert readiness.json()["ready"] is True
    assert readiness.json()["readiness_percent"] == 100
    assert readiness.json()["requirements_pending"] == 1
    assert readiness.json()["crosswalk_verified"] == 0
    assert readiness.json()["unverified"] == 1
    assert readiness.json()["blocking_reasons"] == []


def test_validation_status_only_does_not_stale_current_crosswalk(
    client: TestClient,
) -> None:
    project = _create_project(client, "Optional Validation")
    project_id = str(project["id"])
    requirement_text = "The offeror shall submit a transition plan."
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        f"SECTION L - INSTRUCTIONS\n{requirement_text}",
        name="rfp.pdf",
    )
    proposal_id = seed_extracted_document(
        client,
        project_id,
        requirement_text,
        name="volume-i.pdf",
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    _classify(client, project_id, proposal_id, "PROPOSAL_VOLUME", volume_name="Volume I")
    _extract(client, project_id)
    assert client.post(f"/api/projects/{project_id}/crosswalk/generate").status_code == 200

    requirement = client.get(f"/api/projects/{project_id}/requirements").json()[0]
    validated = client.patch(
        f"/api/projects/{project_id}/requirements/{requirement['id']}",
        json={
            "validation_status": "VALIDATED",
            "reviewer": "Optional Reviewer",
            "expected_updated_at": requirement["updated_at"],
        },
    )
    assert validated.status_code == 200, validated.text
    finding = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    assert finding["stale"] is False
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["ready"] is True
    assert readiness["requirements_validated"] == 1
    assert not any("Regenerate" in reason for reason in readiness["blocking_reasons"])


def test_dismiss_and_reopen_preserve_audit_and_refresh_the_active_crosswalk(
    client: TestClient,
) -> None:
    ready = _ready_project(client)
    project_id = str(ready["project_id"])
    requirement = ready["requirement"]
    assert isinstance(requirement, dict)
    requirement_url = f"/api/projects/{project_id}/requirements/{requirement['id']}"

    dismissed = client.patch(
        requirement_url,
        json={
            "validation_status": "DISMISSED",
            "reviewer": "Exception Reviewer",
            "review_note": "Not applicable to this proposal.",
            "expected_updated_at": requirement["updated_at"],
        },
    )
    assert dismissed.status_code == 200, dismissed.text
    assert client.get(f"/api/projects/{project_id}/crosswalk").json() == []
    assert client.get(f"/api/projects/{project_id}/readiness").json()["requirements_total"] == 0

    reopened = client.patch(
        requirement_url,
        json={
            "validation_status": "PENDING",
            "reviewer": "Exception Reviewer",
            "expected_updated_at": dismissed.json()["updated_at"],
        },
    )
    assert reopened.status_code == 200, reopened.text
    before_regeneration = client.get(f"/api/projects/{project_id}/readiness").json()
    assert any("stale crosswalk" in reason for reason in before_regeneration["blocking_reasons"])

    assert client.post(f"/api/projects/{project_id}/crosswalk/generate").status_code == 200
    refreshed = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    assert refreshed["stale"] is False
    assert refreshed["human_verified"] is False
    after_regeneration = client.get(f"/api/projects/{project_id}/readiness").json()
    assert after_regeneration["ready"] is True
    assert after_regeneration["blocking_reasons"] == []
    reviews = client.get(f"{requirement_url}/reviews").json()
    assert [review["action"] for review in reviews] == ["VALIDATED", "DISMISSED", "REOPENED"]


def test_requirement_change_stales_crosswalk_and_blocks_readiness(client: TestClient) -> None:
    ready = _ready_project(client)
    project_id = str(ready["project_id"])
    requirement = ready["requirement"]
    assert isinstance(requirement, dict)

    response = client.patch(
        f"/api/projects/{project_id}/requirements/{requirement['id']}",
        json={
            "requirement_text": "The contractor shall deliver twenty monthly status reports.",
            "reviewer": "Requirements Reviewer",
            "expected_updated_at": requirement["updated_at"],
        },
    )
    assert response.status_code == 200, response.text
    finding = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    assert finding["stale"] is True
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["ready"] is False
    assert readiness["crosswalk_verified"] == 0
    assert any("Regenerate" in reason for reason in readiness["blocking_reasons"])


def test_manual_crosswalk_override_requires_human_verification(
    client: TestClient,
) -> None:
    ready = _ready_project(client)
    project_id = str(ready["project_id"])
    finding = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    unverified = client.patch(
        f"/api/projects/{project_id}/crosswalk/{finding['id']}",
        json={"human_verified": False},
    )
    assert unverified.status_code == 200, unverified.text

    rejected = client.patch(
        f"/api/projects/{project_id}/crosswalk/{finding['id']}",
        json={"status": "PARTIAL", "reviewer": "Coverage Reviewer"},
    )
    assert rejected.status_code == 422
    assert "Manual status overrides" in rejected.text

    accepted = client.patch(
        f"/api/projects/{project_id}/crosswalk/{finding['id']}",
        json={
            "status": "PARTIAL",
            "human_verified": True,
            "reviewer": "Coverage Reviewer",
        },
    )
    assert accepted.status_code == 200, accepted.text
    assert accepted.json()["needs_attention"] is True
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["ready"] is False
    assert any("coverage gap" in reason for reason in readiness["blocking_reasons"])

    with client.app.state.session_factory() as session:
        proposal = session.scalar(select(Document).where(Document.id == ready["proposal_id"]))
        assert proposal is not None
        proposal.extracted_text = "Unrelated synthetic proposal text."
        proposal.extraction_count = len(proposal.extracted_text)
        session.commit()

    rerun = client.post(f"/api/projects/{project_id}/crosswalk/generate")
    assert rerun.status_code == 200
    assert rerun.json()["verified_findings_marked_stale"] == 1
    stale_override = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    assert stale_override["candidate_status"] == "MISSING"
    assert stale_override["status"] == "PARTIAL"
    assert stale_override["human_verified"] is True
    assert stale_override["stale"] is True
    assert stale_override["needs_attention"] is True
    assert any("Reanalyze" in reason for reason in stale_override["attention_reasons"])


def test_actions_and_saved_workflow_state_do_not_gate_automated_readiness(
    client: TestClient,
) -> None:
    ready = _ready_project(client)
    project_id = str(ready["project_id"])
    action = client.post(
        f"/api/projects/{project_id}/actions",
        json={"title": "Optional follow-up"},
    )
    assert action.status_code == 201, action.text
    blocked_action = client.patch(
        f"/api/projects/{project_id}/actions/{action.json()['id']}",
        json={"status": "BLOCKED"},
    )
    assert blocked_action.status_code == 200, blocked_action.text
    saved_workflow = client.patch(
        f"/api/projects/{project_id}/workflow",
        json={
            "stage": "VERIFY_PACKAGE",
            "status": "BLOCKED",
            "blocker_summary": "Legacy manual blocker",
        },
    )
    assert saved_workflow.status_code == 200, saved_workflow.text

    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["ready"] is True
    assert readiness["actions_blocked"] == 1
    assert readiness["workflow_stage"] == "CROSSWALK"
    assert readiness["workflow_status"] == "COMPLETE"
    assert readiness["blocking_reasons"] == []


def test_solicitation_inventory_change_resets_verification_and_requires_extraction(
    client: TestClient,
) -> None:
    ready = _ready_project(client)
    project_id = str(ready["project_id"])
    amendment_id = seed_extracted_document(
        client,
        project_id,
        "SECTION C\nThe contractor shall submit a new safety plan.",
        name="amendment.pdf",
    )
    _classify(client, project_id, amendment_id, "AMENDMENT")

    checks = client.get(f"/api/projects/{project_id}/intake-verifications").json()
    assert checks
    assert {check["status"] for check in checks} == {"PENDING"}
    assert {check["reviewer"] for check in checks} == {None}
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["ready"] is False
    assert any("new or reclassified" in reason for reason in readiness["blocking_reasons"])
    assert not any("verification checklist" in reason for reason in readiness["blocking_reasons"])


def test_reference_document_is_excluded_from_requirement_extraction_and_readiness(
    client: TestClient,
) -> None:
    project = _create_project(client, "Reference-only project")
    project_id = str(project["id"])
    reference_id = seed_extracted_document(
        client,
        project_id,
        "The contractor shall submit a reference-only report.",
        name="background-reference.pdf",
    )
    _classify(client, project_id, reference_id, "REFERENCE")

    summary = _extract(client, project_id)
    assert summary["documents_analyzed"] == 0
    assert summary["requirements_created"] == 0
    assert client.get(f"/api/projects/{project_id}/requirements").json() == []

    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["requirements_total"] == 0
    assert any(
        "Upload at least one solicitation document" in reason
        for reason in readiness["blocking_reasons"]
    )
    assert not any(
        "new or reclassified solicitation" in reason for reason in readiness["blocking_reasons"]
    )


def test_adding_reference_context_does_not_invalidate_ready_project(client: TestClient) -> None:
    ready = _ready_project(client)
    project_id = str(ready["project_id"])
    reference_id = seed_extracted_document(
        client,
        project_id,
        "Background market research that is context only.",
        name="market-research.pdf",
    )
    _classify(client, project_id, reference_id, "REFERENCE")

    checks = client.get(f"/api/projects/{project_id}/intake-verifications").json()
    assert {check["status"] for check in checks} == {"VERIFIED"}
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["requirements_total"] == 1
    assert readiness["ready"] is True


def test_reclassifying_solicitation_as_reference_retires_active_outputs(
    client: TestClient,
) -> None:
    ready = _ready_project(client)
    project_id = str(ready["project_id"])
    requirement = ready["requirement"]
    assert isinstance(requirement, dict)

    _classify(client, project_id, str(ready["solicitation_id"]), "REFERENCE")

    assert client.get(f"/api/projects/{project_id}/requirements").json() == []
    assert client.get(f"/api/projects/{project_id}/cdrls").json() == []
    assert client.get(f"/api/projects/{project_id}/crosswalk").json() == []
    assert client.get(f"/api/projects/{project_id}/exports/requirements?format=json").json() == []
    assert (
        client.patch(
            f"/api/projects/{project_id}/requirements/{requirement['id']}",
            json={
                "validation_status": "VALIDATED",
                "reviewer": "Should not be reachable",
                "expected_updated_at": requirement["updated_at"],
            },
        ).status_code
        == 404
    )
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["requirements_total"] == 0
    assert readiness["crosswalk_total"] == 0
    assert readiness["ready"] is False
    assert any(
        "Upload at least one solicitation document" in reason
        for reason in readiness["blocking_reasons"]
    )


def test_reclassified_proposal_evidence_is_invalid_even_with_another_volume(
    client: TestClient,
) -> None:
    ready = _ready_project(client, include_backup_proposal=True)
    project_id = str(ready["project_id"])
    finding = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    cited_document_id = finding["evidence"][0]["document_id"]
    assert cited_document_id in {ready["proposal_id"], ready["backup_proposal_id"]}

    _classify(client, project_id, cited_document_id, "REFERENCE")
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["proposal_documents"] == 1
    assert readiness["ready"] is False
    assert readiness["crosswalk_verified"] == 0
    assert any("invalid proposal evidence" in reason for reason in readiness["blocking_reasons"])
    assert any("Regenerate" in reason for reason in readiness["blocking_reasons"])


def test_complete_cdrl_adjudication_is_optional_and_project_scoped(
    client: TestClient,
) -> None:
    project = _create_project(client)
    project_id = str(project["id"])
    document_id, cdrl = _extract_one_cdrl(client, project_id, complete=True)
    cdrl_id = str(cdrl["id"])

    pending = client.get(f"/api/projects/{project_id}/cdrl-adjudications")
    assert pending.status_code == 200
    assert pending.json() == [
        {
            "cdrl_id": cdrl_id,
            "project_id": project_id,
            "status": "PENDING",
            "reviewer": None,
            "waiver_reason": None,
            "reviewed_at": None,
            "updated_at": None,
            "source_fingerprint": None,
            "fresh": False,
            "context_only": False,
            "incomplete": False,
            "missing_fields": [],
            "effective_ready": False,
        }
    ]
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["cdrls_total"] == 1
    assert readiness["cdrls_ready"] == 0
    assert readiness["cdrls_unreviewed"] == 1
    assert not any("CDRL" in reason for reason in readiness["blocking_reasons"])

    url = f"/api/projects/{project_id}/cdrls/{cdrl_id}/adjudication"
    assert client.put(url, json={"status": "REVIEWED"}).status_code == 422
    reviewed = client.put(
        url,
        json={"status": "REVIEWED", "reviewer": "CDRL Reviewer"},
    )
    assert reviewed.status_code == 200, reviewed.text
    assert reviewed.json()["fresh"] is True
    assert reviewed.json()["effective_ready"] is True
    reviewed_at = reviewed.json()["updated_at"]
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["cdrls_ready"] == 1
    assert not any("CDRL" in reason for reason in readiness["blocking_reasons"])

    with client.app.state.session_factory() as session:
        stored = session.scalar(select(CDRL).where(CDRL.id == cdrl_id))
        assert stored is not None
        stored.block_2_title = "CHANGED SYNTHETIC TITLE"
        session.commit()
    stale = client.get(f"/api/projects/{project_id}/cdrl-adjudications").json()[0]
    assert stale["fresh"] is False
    assert stale["effective_ready"] is False
    stale_readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert stale_readiness["cdrls_stale"] == 1
    assert not any("CDRL" in reason for reason in stale_readiness["blocking_reasons"])

    refreshed = client.put(
        url,
        json={
            "status": "REVIEWED",
            "reviewer": "CDRL Reviewer",
            "expected_updated_at": reviewed_at,
        },
    )
    assert refreshed.status_code == 200
    assert refreshed.json()["fresh"] is True
    assert (
        client.put(
            url,
            json={
                "status": "REVIEWED",
                "reviewer": "Stale Reviewer",
                "expected_updated_at": reviewed_at,
            },
        ).status_code
        == 409
    )

    other = _create_project(client, "Other CDRL Project")
    assert client.get(f"/api/projects/{other['id']}/cdrl-adjudications").json() == []
    assert (
        client.put(
            f"/api/projects/{other['id']}/cdrls/{cdrl_id}/adjudication",
            json={"status": "REVIEWED", "reviewer": "Wrong Project"},
        ).status_code
        == 404
    )

    _classify(client, project_id, document_id, "REFERENCE")
    context = client.get(f"/api/projects/{project_id}/cdrl-adjudications").json()[0]
    assert context["context_only"] is True
    assert context["fresh"] is False
    assert context["effective_ready"] is True
    assert client.get(f"/api/projects/{project_id}/readiness").json()["cdrls_total"] == 0

    _classify(client, project_id, document_id, "CDRL")
    reactivated = client.get(f"/api/projects/{project_id}/cdrl-adjudications").json()[0]
    assert reactivated["context_only"] is False
    assert reactivated["fresh"] is False
    assert reactivated["effective_ready"] is False

    reset = client.put(url, json={"status": "PENDING"})
    assert reset.status_code == 200
    assert reset.json()["effective_ready"] is False
    _classify(client, project_id, document_id, "REFERENCE")
    pending_context = client.get(f"/api/projects/{project_id}/cdrl-adjudications").json()[0]
    assert pending_context["context_only"] is True
    assert pending_context["effective_ready"] is True
    assert client.get(f"/api/projects/{project_id}/readiness").json()["cdrls_total"] == 0


def test_optional_incomplete_cdrl_waiver_requires_a_reason(client: TestClient) -> None:
    project = _create_project(client)
    project_id = str(project["id"])
    _, cdrl = _extract_one_cdrl(client, project_id, complete=False)
    cdrl_id = str(cdrl["id"])
    url = f"/api/projects/{project_id}/cdrls/{cdrl_id}/adjudication"

    reviewed = client.put(
        url,
        json={"status": "REVIEWED", "reviewer": "CDRL Reviewer"},
    )
    assert reviewed.status_code == 200
    assert reviewed.json()["incomplete"] is True
    assert reviewed.json()["effective_ready"] is False
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["cdrls_incomplete"] == 1
    assert readiness["cdrls_ready"] == 0
    assert not any("reviewer waiver" in reason for reason in readiness["blocking_reasons"])

    assert (
        client.put(
            url,
            json={"status": "WAIVED", "reviewer": "CDRL Reviewer"},
        ).status_code
        == 422
    )
    waived = client.put(
        url,
        json={
            "status": "WAIVED",
            "reviewer": "Contracts Lead",
            "waiver_reason": "Synthetic form is intentionally partial for this public demo.",
        },
    )
    assert waived.status_code == 200
    assert waived.json()["status"] == "WAIVED"
    assert waived.json()["effective_ready"] is True
    readiness = client.get(f"/api/projects/{project_id}/readiness").json()
    assert readiness["cdrls_ready"] == 1
    assert readiness["cdrls_waived"] == 1
    assert not any("explicit reviewer waiver" in reason for reason in readiness["blocking_reasons"])


def test_manual_crosswalk_evidence_can_be_deleted_but_automated_evidence_cannot(
    client: TestClient,
) -> None:
    project = _create_project(client)
    project_id = str(project["id"])
    requirement_text = "The contractor shall provide a synthetic safety plan."
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        f"SECTION C\n{requirement_text}",
        name="solicitation.pdf",
    )
    proposal_text = f"Introduction. {requirement_text} Conclusion."
    proposal_id = seed_extracted_document(
        client,
        project_id,
        proposal_text,
        name="proposal.pdf",
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    _classify(client, project_id, proposal_id, "PROPOSAL_VOLUME", volume_name="Technical")
    _extract(client, project_id)
    assert client.post(f"/api/projects/{project_id}/crosswalk/generate").status_code == 200
    finding = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    automated = next(item for item in finding["evidence"] if not item["is_manual"])
    automated_url = (
        f"/api/projects/{project_id}/crosswalk/{finding['id']}/evidence/{automated['id']}"
    )
    assert client.delete(automated_url).status_code == 409

    start = proposal_text.index(requirement_text)
    manual = client.post(
        f"/api/projects/{project_id}/crosswalk/{finding['id']}/evidence",
        json={
            "document_id": proposal_id,
            "source_start": start,
            "source_end": start + len(requirement_text),
        },
    )
    assert manual.status_code == 201
    manual_id = manual.json()["id"]
    verified = client.patch(
        f"/api/projects/{project_id}/crosswalk/{finding['id']}",
        json={"human_verified": True, "reviewer": "Evidence Reviewer"},
    )
    assert verified.status_code == 200

    other = _create_project(client, "Other Evidence Project")
    assert (
        client.delete(
            f"/api/projects/{other['id']}/crosswalk/{finding['id']}/evidence/{manual_id}"
        ).status_code
        == 404
    )
    manual_url = f"/api/projects/{project_id}/crosswalk/{finding['id']}/evidence/{manual_id}"
    deleted = client.delete(manual_url)
    assert deleted.status_code == 204
    assert deleted.content == b""
    current = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    assert current["stale"] is True
    assert all(item["id"] != manual_id for item in current["evidence"])
    assert client.delete(manual_url).status_code == 404

    assert client.post(f"/api/projects/{project_id}/crosswalk/generate").status_code == 200
    regenerated = client.get(f"/api/projects/{project_id}/crosswalk").json()[0]
    assert regenerated["stale"] is False
    assert regenerated["human_verified"] is False
    assert all(not item["is_manual"] for item in regenerated["evidence"])


def test_json_csv_and_real_xlsx_exports(client: TestClient) -> None:
    project = _create_project(client)
    project_id = str(project["id"])
    solicitation_id = seed_extracted_document(
        client,
        project_id,
        "SECTION L - INSTRUCTIONS\nThe offeror shall submit Volume I.",
        name="rfp.pdf",
    )
    _classify(client, project_id, solicitation_id, "BASE_SOLICITATION")
    _extract(client, project_id)
    requirement = client.get(f"/api/projects/{project_id}/requirements").json()[0]
    formula_text = '=HYPERLINK("https://invalid.example","unsafe")'
    patched = client.patch(
        f"/api/projects/{project_id}/requirements/{requirement['id']}",
        json={
            "requirement_text": formula_text,
            "reviewer": "Export Safety Test",
            "expected_updated_at": requirement["updated_at"],
        },
    )
    assert patched.status_code == 200

    json_export = client.get(f"/api/projects/{project_id}/exports/requirements?format=json")
    assert json_export.status_code == 200
    assert json_export.headers["content-disposition"].endswith('"')
    assert json_export.json()[0]["source_document"] == "rfp.pdf"
    csv_export = client.get(f"/api/projects/{project_id}/exports/section-l?format=csv")
    assert csv_export.status_code == 200
    assert csv_export.content.startswith(b"\xef\xbb\xbf")
    assert b"source_document" in csv_export.content
    assert b"'=HYPERLINK" in csv_export.content
    assert client.get(f"/api/projects/{project_id}/exports/not-real").status_code == 404

    xlsx = client.get(f"/api/projects/{project_id}/exports/workbook.xlsx")
    assert xlsx.status_code == 200
    assert xlsx.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(io.BytesIO(xlsx.content), read_only=True)
    assert workbook.sheetnames == [
        "Requirements",
        "Section L",
        "Section M",
        "CDRLs",
        "Crosswalk",
        "Readiness",
    ]
    assert workbook["Requirements"]["A1"].value == "id"
    headers = [cell.value for cell in workbook["Requirements"][1]]
    requirement_column = headers.index("requirement") + 1
    assert workbook["Requirements"].cell(2, requirement_column).value.startswith("'=HYPERLINK")
    assert workbook["Readiness"]["A2"].value == project_id
    workbook.close()


def test_new_tables_are_safe_when_reopening_an_existing_sqlite_database(tmp_path: Path) -> None:
    settings = Settings(
        data_dir=tmp_path / "persistent",
        host="127.0.0.1",
        port=8000,
        allowed_origins=("http://127.0.0.1:8000",),
    )
    engine, session_factory = create_database(settings.database_path)
    initialize_database(engine)
    with session_factory() as session:
        project = Project(name="Legacy Project")
        blob = Blob(
            sha256="a" * 64,
            size_bytes=6,
            storage_path="legacy/private-path-that-must-not-be-returned",
        )
        document = Document(
            project=project,
            blob=blob,
            name="legacy.pdf",
            relative_path="legacy.pdf",
            content_type="application/pdf",
            status=DocumentStatus.EXTRACTED,
            extraction_count=6,
            extracted_text="legacy",
        )
        session.add_all((project, blob, document))
        session.commit()
        project_id = project.id
        document_id = document.id
    # Simulate the exact pre-feature state: unchanged core tables and no additive tables.
    with engine.begin() as connection:
        for table_name in (
            "crosswalk_run_states",
            "requirement_extraction_states",
            "project_actions",
            "proposal_evidence",
            "crosswalk_findings",
            "cdrl_adjudications",
            "intake_verifications",
            "document_profiles",
            "project_workflows",
        ):
            connection.exec_driver_sql(f"DROP TABLE {table_name}")
    engine.dispose()

    with make_test_client(create_app(settings)) as reopened:
        projects = reopened.get("/api/projects")
        assert projects.status_code == 200
        assert [item["id"] for item in projects.json()] == [project_id]
        legacy_document = reopened.get(f"/api/projects/{project_id}/documents").json()[0]
        assert legacy_document["id"] == document_id
        assert legacy_document["classification"] == "UNCLASSIFIED"
        assert "private-path" not in str(legacy_document)
        assert _classify(reopened, project_id, document_id, "REFERENCE")["classification"] == (
            "REFERENCE"
        )
        assert reopened.get(f"/api/projects/{project_id}/workflow").status_code == 200
        assert reopened.get(f"/api/projects/{project_id}/readiness").status_code == 200
